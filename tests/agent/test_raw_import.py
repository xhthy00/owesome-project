"""原始成绩导入：匿名编码、班级解析、角色校验、宽表解析。"""

from __future__ import annotations

import inspect
import io
import zipfile
from contextlib import contextmanager
from types import SimpleNamespace
from urllib.parse import unquote

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from common.router import register_routers
from datasource.service.edu_permission import EduScope
from src.agent.education.raw_import import (
    RawDetailRow,
    RawOverviewRow,
    _build_detail_write_rows,
    _encode_school_token,
    _generate_anon_stu_id,
    _ksh_to_class_name,
    _load_overview_dimensions,
    _normalize_id_str,
    _parse_detail_excel,
    _parse_overview_excel,
    _validate_detail_rows,
    _validate_overview_rows,
    assert_raw_import_role_allowed,
    build_raw_detail_template,
    build_raw_overview_template,
    execute_raw_detail_import,
    execute_raw_overview_import,
    preview_raw_detail_import,
    preview_raw_overview_import,
)

_SUBJECTS = ("语文", "数学", "英语", "物理", "化学", "生物", "历史", "政治", "地理")
_OVERVIEW_COL_ORDER = [
    "KSH",
    "SFZH",
    "XM",
    "XX",
    "YW",
    "SX",
    "YY",
    "WL",
    "HX",
    "SW",
    "LS",
    "ZZ",
    "DL",
    "ZF3M",
    "ZF4M",
    "ZF6M",
]


def _overview_row(**overrides):
    base = {
        "KSH": "501101360479",
        "SFZH": "261081010844",
        "XM": "张三",
        "XX": "A05仪征中学",
        "YW": 110,
        "SX": 120,
        "YY": 90,
        "WL": 80,
        "HX": 70,
        "SW": 60,
        "LS": None,
        "ZZ": None,
        "DL": None,
        "ZF3M": 320,
        "ZF4M": 400,
        "ZF6M": 530,
    }
    base.update(overrides)
    return base


def _build_overview_workbook(rows: list[dict], *, omit: set[str] | None = None) -> bytes:
    omit = omit or set()
    headers = [h for h in _OVERVIEW_COL_ORDER if h not in omit]
    extra: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers and key not in omit and key not in extra:
                extra.append(key)
    headers.extend(extra)
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _exam(subject: str, exam_score: float = 150.0) -> dict:
    return {
        "id": f"ex-{subject}",
        "exam_name": f"{subject}卷",
        "subject_name": subject,
        "subject": subject,
        "exam_score": exam_score,
        "exam_time": "2026-01-15",
    }


def _full_exams_by_subject() -> dict[str, dict]:
    return {s: _exam(s) for s in _SUBJECTS}


def _overview_dims(*, schools=None, exams=None, duplicate_subjects=None, batch=None):
    return {
        "batch": batch or {"id": 1, "batch_name": "2026届高三1月期末", "exam_time": "2026-01-15"},
        "schools_by_name": schools if schools is not None else {"A05仪征中学": "GZ_F57E7326"},
        "exams_by_subject": exams if exams is not None else _full_exams_by_subject(),
        "duplicate_subjects": list(duplicate_subjects or []),
    }


def _raw_row(**overrides) -> RawOverviewRow:
    scores = {s: 80.0 for s in _SUBJECTS}
    scores["历史"] = None
    scores["政治"] = None
    scores["地理"] = None
    data = dict(
        row_num=2,
        ksh="501101360479",
        sfzh="261081010844",
        xm="张三",
        xx="A05仪征中学",
        scores=scores,
        totals={"ZF3M": 320.0, "ZF4M": 400.0, "ZF6M": 530.0},
    )
    data.update(overrides)
    return RawOverviewRow(**data)


def test_school_token_matches_golden_pair():
    assert _encode_school_token("A05仪征中学") == "GZ_F57E7326"


def test_anon_stu_id_matches_golden_pair():
    assert _generate_anon_stu_id("GZ_F57E7326", "261081010844") == "GZ_F57E7326_54558B0F"


def test_ksh_to_class_name():
    assert _ksh_to_class_name("501101360479", "2026届高三1月期末") == "高三(10)班"
    # 计划里写的 501101360408[3:5] 实为 "10"；用 08 班号覆盖去前导零
    assert _ksh_to_class_name("501081360408", "2026届高三1月期末") == "高三(8)班"
    assert _ksh_to_class_name("1234", "2026届高三1月期末") == ""


def test_normalize_id_str_strips_excel_float():
    assert _normalize_id_str("261081010844.0") == "261081010844"


def test_resolve_edu_datasource_id_picks_edu_database(monkeypatch):
    from src.agent.education.raw_import import resolve_edu_datasource_id

    rows = [
        SimpleNamespace(id=1, configuration="enc-other"),
        SimpleNamespace(id=3, configuration="enc-edu-b"),
        SimpleNamespace(id=2, configuration="enc-edu-a"),
    ]
    monkeypatch.setattr(
        "datasource.crud.crud_datasource.get_datasources",
        lambda *_a, **_k: rows,
    )
    monkeypatch.setattr(
        "common.utils.aes.decrypt_conf",
        lambda s: {"database": "edu"} if "edu" in str(s) else {"database": "awesome"},
    )
    assert resolve_edu_datasource_id(SimpleNamespace(), 1) == 2


def test_resolve_edu_datasource_id_missing_returns_none(monkeypatch):
    from src.agent.education.raw_import import resolve_edu_datasource_id

    monkeypatch.setattr(
        "datasource.crud.crud_datasource.get_datasources",
        lambda *_a, **_k: [SimpleNamespace(id=1, configuration="enc-other")],
    )
    monkeypatch.setattr("common.utils.aes.decrypt_conf", lambda *_a, **_k: {"database": "awesome"})
    assert resolve_edu_datasource_id(SimpleNamespace(), 1) is None


def test_raw_import_role_rejects_teacher():
    from datasource.service.edu_permission import EduScope

    teacher_err = assert_raw_import_role_allowed(EduScope(edu_role="teacher"))
    assert isinstance(teacher_err, str)
    assert teacher_err
    assert assert_raw_import_role_allowed(EduScope(edu_role="bureau_admin")) is None
    assert assert_raw_import_role_allowed(EduScope(edu_role="school_admin")) is None


def test_parse_overview_basic_two_rows():
    data = _build_overview_workbook([
        _overview_row(),
        _overview_row(KSH="501081360408", SFZH="261081010845", XM="李四"),
    ])
    rows, errors, _headers = _parse_overview_excel(data)
    assert errors == []
    assert len(rows) == 2
    assert rows[0].row_num == 2
    assert rows[1].row_num == 3
    assert rows[0].ksh == "501101360479"
    assert rows[0].sfzh == "261081010844"
    assert rows[0].xm == "张三"
    assert rows[0].xx == "A05仪征中学"
    assert rows[0].scores["语文"] == 110.0
    assert rows[0].scores["历史"] is None
    assert rows[0].totals["ZF6M"] == 530.0
    assert rows[1].ksh == "501081360408"
    assert rows[1].xm == "李四"


def test_parse_overview_keeps_passthrough_others():
    data = _build_overview_workbook([
        _overview_row(XKKM="物理化学生物", XSXZ="应届"),
    ])
    rows, errors, _headers = _parse_overview_excel(data)
    assert errors == []
    assert rows[0].others["XKKM"] == "物理化学生物"
    assert rows[0].others["XSXZ"] == "应届"
    assert "KSH" not in rows[0].others
    assert "YW" not in rows[0].others
    assert "ZF6M" not in rows[0].others


def test_parse_overview_missing_required_columns():
    data = _build_overview_workbook([_overview_row()], omit={"ZF6M"})
    rows, errors, _headers = _parse_overview_excel(data)
    assert rows == []
    assert len(errors) == 1
    assert errors[0].row == 0
    assert errors[0].field == "header"
    assert "ZF6M" in errors[0].message


def test_parse_overview_sfzh_duplicate():
    data = _build_overview_workbook([
        _overview_row(),
        _overview_row(KSH="501081360408", SFZH="261081010844", XM="李四"),
    ])
    rows, errors, _headers = _parse_overview_excel(data)
    assert len(rows) == 1
    assert rows[0].ksh == "501101360479"
    assert len(errors) == 1
    assert errors[0].row == 3
    assert errors[0].field == "SFZH"


def test_parse_overview_ksh_duplicate():
    data = _build_overview_workbook([
        _overview_row(),
        _overview_row(KSH="501101360479", SFZH="261081010845", XM="李四"),
    ])
    rows, errors, _headers = _parse_overview_excel(data)
    assert len(rows) == 1
    assert rows[0].sfzh == "261081010844"
    assert len(errors) == 1
    assert errors[0].row == 3
    assert errors[0].field == "KSH"


def test_parse_overview_corrupt_file_is_whole_file_error():
    rows, errors, _headers = _parse_overview_excel(b"not-an-excel-file")
    assert rows == []
    assert len(errors) == 1
    assert errors[0].row == 0
    assert errors[0].field in {"header", "file"}
    assert "无法解析" in errors[0].message


def test_validate_overview_school_missing():
    row = _raw_row(xx="不存在的学校")
    valid, errors = _validate_overview_rows([row], _overview_dims())
    assert valid == []
    assert any("学校『不存在的学校』不存在" in e.message for e in errors)


def test_validate_overview_score_out_of_range():
    scores = {s: 80.0 for s in _SUBJECTS}
    scores["语文"] = 200.0
    row = _raw_row(scores=scores)
    valid, errors = _validate_overview_rows([row], _overview_dims())
    assert valid == []
    assert any("超满分" in e.message for e in errors)


def test_validate_overview_missing_required_paper():
    exams = _full_exams_by_subject()
    del exams["物理"]
    valid, errors = _validate_overview_rows([_raw_row()], _overview_dims(exams=exams))
    assert valid == []
    assert errors[0].row == 0
    assert "物理" in errors[0].message
    assert "缺少" in errors[0].message


def test_validate_overview_duplicate_subject_papers():
    valid, errors = _validate_overview_rows(
        [_raw_row()],
        _overview_dims(duplicate_subjects=["物理"]),
    )
    assert valid == []
    assert errors[0].row == 0
    assert "物理" in errors[0].message


def test_validate_overview_school_admin_rejects_foreign_school():
    own = _raw_row()
    other = _raw_row(
        row_num=3,
        ksh="501081360408",
        sfzh="261081010845",
        xm="李四",
        xx="A01扬州中学",
    )
    dims = _overview_dims(
        schools={"A05仪征中学": "GZ_F57E7326", "A01扬州中学": "GZ_OTHER001"},
    )
    scope = EduScope(edu_role="school_admin", school_id="GZ_F57E7326")
    valid, errors = _validate_overview_rows([own, other], dims, scope=scope)
    assert valid == []
    assert any(e.row == 0 and "非本校" in e.message and "A01扬州中学" in e.message for e in errors)

    bureau, bureau_errors = _validate_overview_rows(
        [own, other],
        dims,
        scope=EduScope(edu_role="bureau_admin"),
    )
    assert len(bureau) == 2
    assert bureau_errors == []


def test_validate_overview_school_admin_requires_school_id():
    valid, errors = _validate_overview_rows(
        [_raw_row()],
        _overview_dims(),
        scope=EduScope(edu_role="school_admin", school_id=""),
    )
    assert valid == []
    assert len(errors) == 1
    assert errors[0].row == 0
    assert "校管理员未配置 school_id" in errors[0].message


def test_load_overview_dimensions_maps_batch_schools_exams(monkeypatch):
    def fake_execute_sql(_db_type, _config, sql):
        sql_l = sql.lower()
        if "tb_exam_batch" in sql_l:
            return True, "ok", {
                "columns": ["id", "batch_name", "exam_time"],
                "rows": [(1, "2026届高三1月期末", "2026-01-15")],
            }
        if "tb_school" in sql_l:
            return True, "ok", {
                "columns": ["id", "s_name"],
                "rows": [("GZ_F57E7326", "A05仪征中学")],
            }
        if "tb_exam" in sql_l:
            return True, "ok", {
                "columns": ["id", "exam_name", "subject_name", "subject", "exam_score", "exam_time"],
                "rows": [
                    (11, "语文卷", "语文", "语文", 150, "2026-01-15"),
                    (12, "语文卷B", "语文", "语文", 150, "2026-01-15"),
                    (13, "数学卷", "数学", "数学", 150, "2026-01-15"),
                ],
            }
        return False, f"unexpected sql: {sql}", None

    monkeypatch.setattr("src.agent.education.raw_import.execute_sql", fake_execute_sql)
    dims = _load_overview_dimensions("pg", {}, 1, {"A05仪征中学"})
    assert dims["batch"]["batch_name"] == "2026届高三1月期末"
    assert dims["schools_by_name"]["A05仪征中学"] == "GZ_F57E7326"
    assert dims["exams_by_subject"]["语文"]["id"] == 11
    assert "语文" in dims["duplicate_subjects"]
    assert "数学" in dims["exams_by_subject"]


def test_load_overview_dimensions_missing_batch_raises(monkeypatch):
    def fake_execute_sql(_db_type, _config, sql):
        if "tb_exam_batch" in sql.lower():
            return True, "ok", {"columns": ["id", "batch_name", "exam_time"], "rows": []}
        return False, "unexpected", None

    monkeypatch.setattr("src.agent.education.raw_import.execute_sql", fake_execute_sql)
    with pytest.raises(ValueError):
        _load_overview_dimensions("pg", {}, 99, set())


_OVERVIEW_SCHEMA_COLS = {
    "ksh",
    "exam_name",
    "exam_batch_id",
    "sfzh",
    "xm",
    "xx",
    "bj",
    "anon_stu_id",
    "yw",
    "sx",
    "yy",
    "wl",
    "hx",
    "sw",
    "ls",
    "zz",
    "dl",
    "zf3m",
    "zf4m",
    "zf6m",
}
_STUDENT_SCHEMA_COLS = {"id", "school_id", "class", "jc"}
_SCORE_SCHEMA_COLS = {
    "exam_id",
    "student_id",
    "school_id",
    "class",
    "score",
    "subject_name",
    "exam_score",
    "exam_time",
}
_PII_KEYS = {"xm", "sfzh", "ksh"}
_PREVIEW_HIDDEN_KEYS = {"ry", "rykg", "ryzw"}


def _two_subject_overview_row(**overrides):
    return _overview_row(
        YW=110,
        SX=120,
        YY=None,
        WL=None,
        HX=None,
        SW=None,
        LS=None,
        ZZ=None,
        DL=None,
        **overrides,
    )


def _patch_overview_io(monkeypatch, *, dims=None):
    dims = dims or _overview_dims()
    monkeypatch.setattr(
        "src.agent.education.raw_import._load_overview_dimensions",
        lambda *_a, **_k: dims,
    )
    calls: list[dict] = []

    def fake_upsert(_db_type, _config, table, cols, conflict_cols, rows):
        calls.append(
            {
                "table": table,
                "cols": list(cols),
                "conflict_cols": tuple(conflict_cols),
                "rows": list(rows),
            }
        )
        return len(rows)

    monkeypatch.setattr("src.agent.education.raw_import._upsert_dict_rows", fake_upsert)

    def fake_schema(_db_type, _config, table):
        name = str(table or "")
        if "overview" in name:
            return set(_OVERVIEW_SCHEMA_COLS)
        if "student" in name:
            return set(_STUDENT_SCHEMA_COLS)
        return set(_SCORE_SCHEMA_COLS)

    monkeypatch.setattr("src.agent.education.raw_import._schema_columns", fake_schema)
    return calls


def test_execute_overview_upserts_three_tables_for_two_subjects(monkeypatch):
    calls = _patch_overview_io(monkeypatch)
    data = _build_overview_workbook([_two_subject_overview_row()])
    result = execute_raw_overview_import(
        data,
        1,
        EduScope(edu_role="bureau_admin"),
        "pg",
        {},
    )
    assert result.error_rows == []
    assert result.valid_rows == 1
    assert len(calls) == 3
    overview_call, student_call, score_call = calls
    assert "overview" in overview_call["table"]
    assert overview_call["conflict_cols"] == ("ksh", "exam_name")
    assert len(overview_call["rows"]) == 1
    ov = overview_call["rows"][0]
    assert ov["ksh"] == "501101360479"
    assert ov["exam_name"] == "2026届高三1月期末"
    assert ov["anon_stu_id"] == "GZ_F57E7326_54558B0F"
    assert ov["bj"] == "高三(10)班"
    assert ov["yw"] == 110.0
    assert ov["sx"] == 120.0
    assert student_call["conflict_cols"] == ("id",)
    assert student_call["rows"][0]["id"] == "GZ_F57E7326_54558B0F"
    assert student_call["rows"][0]["school_id"] == "GZ_F57E7326"
    assert student_call["rows"][0]["class"] == "高三(10)班"
    assert student_call["rows"][0]["jc"] == "2026"
    assert score_call["conflict_cols"] == ("exam_id", "student_id")
    assert len(score_call["rows"]) == 2
    subjects = {r["subject_name"] for r in score_call["rows"]}
    assert subjects == {"语文", "数学"}
    assert result.summary["overview_upserted"] == 1
    assert result.summary["students_upserted"] == 1
    assert result.summary["score_upserted"] == 2
    assert all(isinstance(v, int) for v in result.summary.values())


def test_execute_overview_upsert_raise_is_write_error(monkeypatch):
    _patch_overview_io(monkeypatch)

    def boom(*_a, **_k):
        raise RuntimeError("connection lost")

    monkeypatch.setattr("src.agent.education.raw_import._upsert_dict_rows", boom)
    data = _build_overview_workbook([_two_subject_overview_row()])
    result = execute_raw_overview_import(
        data,
        1,
        EduScope(edu_role="bureau_admin"),
        "pg",
        {},
    )
    assert result.error_rows
    assert result.error_rows[0].field == "写入"
    assert "connection lost" in result.error_rows[0].message


def test_execute_overview_skips_write_when_school_missing(monkeypatch):
    calls = _patch_overview_io(monkeypatch)
    data = _build_overview_workbook([_two_subject_overview_row(XX="不存在的学校")])
    result = execute_raw_overview_import(
        data,
        1,
        EduScope(edu_role="bureau_admin"),
        "pg",
        {},
    )
    assert result.error_rows
    assert any("不存在" in e.message for e in result.error_rows)
    assert calls == []
    assert result.summary == {} or result.summary.get("overview_upserted", 0) == 0


def test_preview_sample_includes_identity_cols(monkeypatch):
    _patch_overview_io(monkeypatch)
    data = _build_overview_workbook([_two_subject_overview_row()])
    result = preview_raw_overview_import(
        data,
        1,
        EduScope(edu_role="bureau_admin"),
        "pg",
        {},
    )
    assert result.error_rows == []
    assert result.valid_rows == 1
    assert len(result.preview_sample) == 1
    sample = result.preview_sample[0]
    for key in _PREVIEW_HIDDEN_KEYS:
        assert key not in sample
        assert key not in {str(k).lower() for k in sample}
    assert sample["anon_stu_id"] == "GZ_F57E7326_54558B0F"
    assert sample["ksh"] == "501101360479"
    assert sample["sfzh"] == "261081010844"
    assert sample["xm"] == "张三"
    assert sample["xx"] == "A05仪征中学"
    assert sample["bj"] == "高三(10)班"
    assert sample["zf6m"] == 530.0
    assert sample["yw"] == 110.0
    titles = {c["key"]: c["title"] for c in result.preview_columns}
    assert titles["anon_stu_id"] == "学号匿名编码"
    assert titles["ksh"] == "考生号"
    assert titles["sfzh"] == "学号"
    assert titles["xm"] == "姓名"
    assert titles["bj"] == "班级"
    assert titles["xx"] == "学校"
    assert titles["zf6m"] == "全科总分"
    assert titles["yw"] == "语文"
    assert [c["key"] for c in result.preview_columns] == list(sample)
    assert result.resolved_rows
    for rec in result.resolved_rows:
        assert rec["school_id"] == "GZ_F57E7326"
        for key in _PII_KEYS:
            assert key not in rec


def test_preview_sample_returns_all_valid_rows(monkeypatch):
    _patch_overview_io(monkeypatch)
    rows = [
        _two_subject_overview_row(
            KSH=f"5011013604{i:02d}",
            SFZH=f"2610810108{i:02d}",
            XM=f"学生{i}",
        )
        for i in range(12)
    ]
    result = preview_raw_overview_import(
        _build_overview_workbook(rows),
        1,
        EduScope(edu_role="bureau_admin"),
        "pg",
        {},
    )
    assert result.error_rows == []
    assert result.valid_rows == 12
    assert len(result.preview_sample) == 12


def test_preview_sample_follows_excel_and_hides_ry_cols(monkeypatch):
    _patch_overview_io(monkeypatch)
    data = _build_overview_workbook(
        [
            _two_subject_overview_row(
                DQ="仪征",
                XKKM="物化生",
                RY=90,
                RYKG=40,
                RYZW=20,
            )
        ]
    )
    result = preview_raw_overview_import(
        data,
        1,
        EduScope(edu_role="bureau_admin"),
        "pg",
        {},
    )
    assert result.error_rows == []
    sample = result.preview_sample[0]
    keys = [c["key"] for c in result.preview_columns]
    assert keys[0] == "anon_stu_id"
    assert "ksh" in keys
    assert "sfzh" in keys
    assert "xm" in keys
    assert "bj" in keys
    assert keys.index("bj") == keys.index("xx") + 1
    assert "dq" in keys
    assert "xkkm" in keys
    assert sample["dq"] == "仪征"
    assert sample["xkkm"] == "物化生"
    assert sample["ksh"] == "501101360479"
    assert sample["sfzh"] == "261081010844"
    assert sample["xm"] == "张三"
    for hidden in ("ry", "rykg", "ryzw"):
        assert hidden not in keys
        assert hidden not in sample
    titles = {c["key"]: c["title"] for c in result.preview_columns}
    assert titles["dq"] == "地区"
    assert titles["xkkm"] == "选考科目组合"


def test_school_admin_foreign_file_execute_does_not_write(monkeypatch):
    dims = _overview_dims(
        schools={"A05仪征中学": "GZ_F57E7326", "A01扬州中学": "GZ_OTHER001"},
    )
    calls = _patch_overview_io(monkeypatch, dims=dims)
    data = _build_overview_workbook(
        [
            _two_subject_overview_row(),
            _two_subject_overview_row(
                KSH="501081360408",
                SFZH="261081010845",
                XM="李四",
                XX="A01扬州中学",
            ),
        ]
    )
    result = execute_raw_overview_import(
        data,
        1,
        EduScope(edu_role="school_admin", school_id="GZ_F57E7326"),
        "pg",
        {},
    )
    assert result.valid_rows == 0
    assert any(e.row == 0 and "非本校" in e.message for e in result.error_rows)
    assert calls == []


def _build_detail_workbook(
    headers: list[str],
    data: list[list],
    *,
    title: str = "小题分(数学)",
    info_row: list | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append([title])
    ws.append(info_row or ["区域占位"])
    ws.append(headers)
    for row in data:
        ws.append(row)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_parse_detail_basic():
    data = _build_detail_workbook(
        ["学号", "考号", "姓名", "学校", "单选1（5.0分）", "12（5.0分）", "15_1（6.0分）"],
        [["261081010844", "501101360479", "张三", "A01扬州中学", 5, 5, 6]],
    )
    rows, questions, errors, subject = _parse_detail_excel(data)
    assert not errors
    assert subject == "数学"
    assert len(rows) == 1
    assert rows[0].scores == {"单选1": 5.0, "12": 5.0, "15_1": 6.0}
    assert rows[0].sfzh == "261081010844"
    assert rows[0].ksh == "501101360479"
    assert rows[0].xm == "张三"
    assert rows[0].school_name == "A01扬州中学"
    assert {q["question_no"] for q in questions} == {"单选1", "12", "15_1"}


def test_parse_detail_skips_answer_and_total_columns():
    data = _build_detail_workbook(
        [
            "学号",
            "单选1（5.0分）",
            "单选1_答案",
            "全卷（150分）",
            "1卷（50分）",
            "15（13.0分）",
            "15_1（6.0分）",
        ],
        [["261081010844", 5, "A", 150, 50, 13, 6]],
    )
    rows, questions, errors, _subject = _parse_detail_excel(data)
    assert not errors
    assert rows[0].scores == {"单选1": 5.0, "15": 13.0, "15_1": 6.0}
    assert "全卷" not in rows[0].scores
    nos = {q["question_no"] for q in questions}
    assert "全卷" not in nos
    assert "15" in nos and "15_1" in nos


def test_parse_detail_empty_score_is_zero():
    data = _build_detail_workbook(
        ["学号", "单选1（5.0分）"],
        [["261081010844", None]],
    )
    rows, _questions, errors, _subject = _parse_detail_excel(data)
    assert not errors
    assert rows[0].scores == {"单选1": 0.0}


def _detail_dims(**overrides):
    dims = {
        "batch": {"id": 1, "batch_name": "2026届高三1月期末"},
        "exam": {"id": 10, "exam_batch_id": 1, "subject_name": "数学"},
        "overview_by_sfzh": {
            "261081010844": {
                "sfzh": "261081010844",
                "anon_stu_id": "GZ_F57E7326_54558B0F",
                "xx": "A05仪征中学",
                "bj": "高三(10)班",
            }
        },
        "questions_by_no": {
            "单选1": {"id": 101, "question_score": 5},
            "12": {"id": 102, "question_score": 5},
            "15_1": {"id": 103, "question_score": 6},
        },
    }
    dims.update(overrides)
    return dims


def test_build_detail_write_rows_uses_exam_question_id():
    resolved = [
        {
            "row": RawDetailRow(
                row_num=4,
                sfzh="261081010844",
                ksh="",
                school_name="",
                scores={"24-2": 4.0},
            ),
            "anon_stu_id": "GZ_E74EBD64_9DCD1578",
            "school_token": "GZ_E74EBD64",
            "class_name": "高三(1)班",
        }
    ]
    rows = _build_detail_write_rows(
        resolved,
        {
            "exam": {"id": 33},
            "questions_by_no": {"24-2": {"id": 88, "question_score": 4.0}},
        },
    )
    assert len(rows) == 1
    assert rows[0]["exam_id"] == 33
    assert rows[0]["question_no"] == "24-2"
    assert rows[0]["question_id"] == 88
    assert rows[0]["question_id"] != ""
    assert rows[0]["score"] == 4.0
    assert rows[0]["question_score"] == 4.0


def test_build_detail_write_rows_blank_question_id_is_none():
    resolved = [
        {
            "row": RawDetailRow(
                row_num=4,
                sfzh="261081010844",
                ksh="",
                school_name="",
                scores={"24-2": 4.0},
            ),
            "anon_stu_id": "GZ_E74EBD64_9DCD1578",
            "school_token": "GZ_E74EBD64",
            "class_name": "高三(1)班",
        }
    ]
    rows = _build_detail_write_rows(
        resolved,
        {"exam": {"id": 33}, "questions_by_no": {"24-2": {"question_score": 4.0}}},
    )
    assert rows[0]["question_id"] is None


def test_execute_detail_upserts_when_overview_hits(monkeypatch):
    calls: list[tuple] = []

    def fake_load(*_a, **_k):
        return _detail_dims()

    def fake_upsert(*args, **_k):
        calls.append(args)
        return args[-1] and len(args[-1])

    monkeypatch.setattr("src.agent.education.raw_import._load_detail_dimensions", fake_load)
    monkeypatch.setattr("src.agent.education.raw_import._upsert_dict_rows", fake_upsert)
    monkeypatch.setattr("src.agent.education.raw_import._schema_columns", lambda *_a, **_k: set())
    data = _build_detail_workbook(
        ["学号", "单选1（5.0分）", "12（5.0分）"],
        [["261081010844", 5, 4]],
    )
    result = execute_raw_detail_import(
        data, 1, 10, EduScope(edu_role="bureau_admin"), "pg", {}
    )
    assert not result.error_rows
    assert result.summary["detail_upserted"] == 2
    assert result.summary["students_matched"] == 1
    assert len(calls) == 1
    written = calls[0][-1]
    assert {row["question_id"] for row in written} == {101, 102}
    assert all(row["question_id"] not in ("",) for row in written)


def test_execute_detail_skips_write_when_student_missing_from_overview(monkeypatch):
    calls: list = []
    monkeypatch.setattr(
        "src.agent.education.raw_import._load_detail_dimensions",
        lambda *_a, **_k: _detail_dims(overview_by_sfzh={}),
    )
    monkeypatch.setattr(
        "src.agent.education.raw_import._upsert_dict_rows",
        lambda *_a, **_k: calls.append(1) or 0,
    )
    data = _build_detail_workbook(["学号", "单选1（5.0分）"], [["261081010844", 5]])
    result = execute_raw_detail_import(
        data, 1, 10, EduScope(edu_role="bureau_admin"), "pg", {}
    )
    assert any("不在已导入的宽表中" in e.message for e in result.error_rows)
    assert calls == []


def test_validate_detail_unknown_question_no_rejects_file():
    data = _build_detail_workbook(["学号", "单选1（5.0分）"], [["261081010844", 5]])
    rows, questions, parse_errors, _subject = _parse_detail_excel(data)
    assert not parse_errors
    valid, errors = _validate_detail_rows(
        rows,
        questions,
        _detail_dims(questions_by_no={"12": {"question_score": 5}}),
        EduScope(edu_role="bureau_admin"),
    )
    assert valid == []
    assert any(e.row == 0 and "题号" in e.message and "不存在" in e.message for e in errors)


def test_preview_detail_sample_includes_excel_and_detail_cols(monkeypatch):
    monkeypatch.setattr(
        "src.agent.education.raw_import._load_detail_dimensions",
        lambda *_a, **_k: _detail_dims(),
    )
    data = _build_detail_workbook(
        ["学号", "考号", "姓名", "学校", "单选1（5.0分）", "12（5.0分）", "15_1（6.0分）"],
        [["261081010844", "501101360479", "张三", "A01扬州中学", 5, 5, 6]],
    )
    result = preview_raw_detail_import(
        data, 1, 10, EduScope(edu_role="bureau_admin"), "pg", {}
    )
    assert not result.error_rows
    assert len(result.preview_sample) == 1
    sample = result.preview_sample[0]
    assert sample["exam_id"] == "10"
    assert sample["student_id"] == "GZ_F57E7326_54558B0F"
    assert sample["anon_stu_id"] == sample["student_id"]
    assert sample["class"] == "高三(10)班"
    assert sample["ksh"] == "501101360479"
    assert sample["sfzh"] == "261081010844"
    assert sample["xm"] == "张三"
    assert sample["xx"] == "A01扬州中学"
    assert sample["单选1"] == 5.0
    assert sample["12"] == 5.0
    assert sample["15_1"] == 6.0
    assert "questions" not in sample
    titles = {c["key"]: c["title"] for c in result.preview_columns}
    assert titles["exam_id"] == "试卷ID"
    assert titles["student_id"] == "学号匿名编码"
    assert titles["class"] == "班级"
    assert titles["ksh"] == "考生号"
    assert titles["sfzh"] == "学号"
    assert titles["xm"] == "姓名"
    assert titles["xx"] == "学校"
    assert titles["单选1"] == "单选1（5.0分）"
    assert [c["key"] for c in result.preview_columns] == [
        "exam_id",
        "student_id",
        "class",
        "ksh",
        "sfzh",
        "xm",
        "xx",
        "单选1",
        "12",
        "15_1",
    ]


def test_preview_detail_sample_returns_all_valid_rows(monkeypatch):
    overview = {
        f"26108101{i:04d}": {
            "sfzh": f"26108101{i:04d}",
            "anon_stu_id": f"GZ_F57E7326_{i:08X}",
            "xx": "A05仪征中学",
            "bj": "高三(10)班",
        }
        for i in range(12)
    }
    monkeypatch.setattr(
        "src.agent.education.raw_import._load_detail_dimensions",
        lambda *_a, **_k: _detail_dims(overview_by_sfzh=overview),
    )
    data = _build_detail_workbook(
        ["学号", "单选1（5.0分）"],
        [[f"26108101{i:04d}", 5] for i in range(12)],
    )
    result = preview_raw_detail_import(
        data, 1, 10, EduScope(edu_role="bureau_admin"), "pg", {}
    )
    assert not result.error_rows
    assert result.valid_rows == 12
    assert len(result.preview_sample) == 12
    assert result.preview_sample[0]["单选1"] == 5.0
    assert result.preview_sample[-1]["sfzh"] == "261081010011"


def test_execute_detail_school_admin_rejects_foreign(monkeypatch):
    calls: list = []
    monkeypatch.setattr(
        "src.agent.education.raw_import._load_detail_dimensions",
        lambda *_a, **_k: _detail_dims(),
    )
    monkeypatch.setattr(
        "src.agent.education.raw_import._upsert_dict_rows",
        lambda *_a, **_k: calls.append(1) or 0,
    )
    data = _build_detail_workbook(["学号", "单选1（5.0分）"], [["261081010844", 5]])
    result = execute_raw_detail_import(
        data, 1, 10, EduScope(edu_role="school_admin", school_id="GZ_OTHER001"), "pg", {}
    )
    assert result.valid_rows == 0
    assert any(e.row == 0 and "非本校" in e.message for e in result.error_rows)
    assert calls == []


def test_scan_alerts_batch_wide_uses_all_exams(monkeypatch):
    from src.agent.education.alert_service import scan_alerts_after_import

    detect_calls: list[tuple[str, str]] = []

    def fake_sql(_db_type, _config, sql):
        assert "exam_batch_id" in sql
        return True, "", {"columns": ["id", "exam_name"], "rows": [[1, "语文卷"], [2, "数学卷"]]}

    def fake_detect(_session, **kwargs):
        detect_calls.append((kwargs["school_id"], kwargs["exam_id"]))
        assert kwargs.get("class_names") == []
        return {"inserted": 1, "updated": 0, "skipped": 0, "detected": 1}

    monkeypatch.setattr("datasource.db.db.execute_sql", fake_sql)
    monkeypatch.setattr(
        "src.agent.education.alert_service.detect_and_upsert_for_exam", fake_detect
    )
    result = scan_alerts_after_import(
        None,
        db_type="pg",
        config={},
        workspace_oid=1,
        datasource_id=12,
        resolved_rows=[{"school_id": "GZ_19D9D68D"}],
        exam_batch_id=1,
    )
    assert len(detect_calls) == 2
    assert result["inserted"] == 2
    assert result["exams"] == 2


def _unauth_client() -> TestClient:
    app = FastAPI()
    register_routers(app)
    return TestClient(app)


def _raw_auth_client(
    monkeypatch, *, edu_role: str = "bureau_admin", edu_ds_id: int | None = 12
) -> TestClient:
    from common.middlewares.exception import register_exception_handlers
    from src.agent.resource.tool import business as biz
    from system.api.auth_deps import get_current_user
    from system.schemas import UserResponse
    from system.workspace_scope import get_workspace_oid

    monkeypatch.setattr(
        "src.agent.education.raw_import.resolve_edu_datasource_id",
        lambda *_a, **_k: edu_ds_id,
    )
    monkeypatch.setattr(
        "src.agent.education.api.assert_datasource_accessible",
        lambda *a, **kw: SimpleNamespace(id=12, oid=1),
    )
    monkeypatch.setattr(biz, "_load_datasource", lambda *a, **kw: ("pg", {}, "ds"))
    monkeypatch.setattr(
        "system.crud.crud_user.get_user_by_id",
        lambda session, uid: SimpleNamespace(
            id=uid,
            system_variables={"edu_role": edu_role, "school_id": "GZ_F57E7326"},
        ),
    )

    @contextmanager
    def fake_db():
        yield SimpleNamespace()

    monkeypatch.setattr("common.core.database.get_db_session", fake_db)
    monkeypatch.setattr("audit.service.decorators._write_access_log", lambda *a, **k: None)

    app = FastAPI()
    register_exception_handlers(app)
    register_routers(app)
    app.dependency_overrides[get_current_user] = lambda: UserResponse(
        id=1,
        account="admin",
        name="Admin",
        email=None,
        oid=1,
        status=1,
        language="zh-CN",
        origin=0,
        create_time=0,
    )
    app.dependency_overrides[get_workspace_oid] = lambda: 1
    return TestClient(app)


def test_raw_import_batches_requires_auth():
    client = _unauth_client()
    r = client.get("/api/v1/education/raw-score-import/batches")
    assert r.status_code in (401, 403)


def test_old_score_import_template_route_still_exists():
    client = _unauth_client()
    r = client.get("/api/v1/education/score-import/templates/total")
    assert r.status_code != 404


def test_raw_import_endpoints_have_no_datasource_id_param():
    from src.agent.education import api as edu_api

    for fn in (
        edu_api.list_raw_import_batches,
        edu_api.create_raw_import_batch,
        edu_api.list_raw_import_papers,
        edu_api.preview_raw_overview,
        edu_api.execute_raw_overview,
        edu_api.preview_raw_detail,
        edu_api.execute_raw_detail,
        edu_api.download_raw_import_template,
    ):
        assert "datasource_id" not in inspect.signature(fn).parameters
    assert "datasource_id" in inspect.signature(edu_api.preview_score_import).parameters
    assert "datasource_id" in inspect.signature(edu_api.execute_score_import).parameters


def test_raw_import_teacher_forbidden(monkeypatch):
    client = _raw_auth_client(monkeypatch, edu_role="teacher")
    r = client.get("/api/v1/education/raw-score-import/batches")
    assert r.status_code == 200
    body = r.json()
    assert body["code"] == 403
    assert body["message"]


def _patch_write_session(monkeypatch, *, existing: dict | None = None, inserted: dict | None = None):
    calls: dict = {"query": [], "write": [], "commit": 0}

    class FakeWriteSession:
        def __init__(self, db_type, config):
            self.db_type = db_type
            self.config = config

        def __enter__(self):
            return self

        def __exit__(self, *_a):
            return False

        def execute_query(self, sql, params=None):
            calls["query"].append((sql, params))
            if existing is not None:
                rec = existing
                return True, "ok", {
                    "columns": ["id", "batch_name", "exam_time"],
                    "rows": [(rec["id"], rec["batch_name"], rec["exam_time"])],
                }
            if len(calls["query"]) == 1:
                return True, "ok", {
                    "columns": ["id", "batch_name", "exam_time"],
                    "rows": [],
                }
            rec = inserted or {}
            return True, "ok", {
                "columns": ["id", "batch_name", "exam_time"],
                "rows": [(rec["id"], rec["batch_name"], rec["exam_time"])],
            }

        def execute_write(self, sql, params=None):
            calls["write"].append((sql, params))
            return True, "ok", {"row_count": 1}

        def commit(self):
            calls["commit"] += 1

        def rollback(self):
            pass

    monkeypatch.setattr("datasource.db.db.WriteDbSession", FakeWriteSession)
    return calls


def test_create_raw_import_batch_success(monkeypatch):
    inserted = {"id": 7, "batch_name": "2026届高三1月期末", "exam_time": "2026-01-15"}
    calls = _patch_write_session(monkeypatch, inserted=inserted)
    client = _raw_auth_client(monkeypatch)
    r = client.post(
        "/api/v1/education/raw-score-import/batches",
        data={"batch_name": "2026届高三1月期末", "exam_time": "2026-01-15"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["code"] == 200
    assert body["data"] == inserted
    assert body["data"]["exam_time"] == "2026-01-15"
    assert len(calls["query"]) >= 1
    lookup_sql, lookup_params = calls["query"][0]
    assert "%s" in lookup_sql
    assert lookup_params == ("2026届高三1月期末",)
    assert "'" not in lookup_sql.split("WHERE", 1)[-1]
    assert calls["write"]
    write_sql, write_params = calls["write"][0]
    assert write_params == ("2026届高三1月期末", "2026-01-15")
    assert "%s" in write_sql
    assert calls["commit"] == 1


def test_create_raw_import_batch_duplicate(monkeypatch):
    existing = {"id": 3, "batch_name": "2026届高三1月期末", "exam_time": "2026-01-01"}
    calls = _patch_write_session(monkeypatch, existing=existing)
    client = _raw_auth_client(monkeypatch)
    r = client.post(
        "/api/v1/education/raw-score-import/batches",
        data={"batch_name": "2026届高三1月期末", "exam_time": "2026-01-15"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["code"] == 400
    assert body["data"]["id"] == 3
    assert body["data"]["batch_name"] == "2026届高三1月期末"
    assert "已存在" in body["message"]
    lookup_sql, lookup_params = calls["query"][0]
    assert "%s" in lookup_sql
    assert lookup_params == ("2026届高三1月期末",)
    assert calls["write"] == []
    assert calls["commit"] == 0


def test_create_raw_import_batch_empty_name(monkeypatch):
    calls = _patch_write_session(monkeypatch)
    client = _raw_auth_client(monkeypatch)
    r = client.post(
        "/api/v1/education/raw-score-import/batches",
        data={"batch_name": "   ", "exam_time": "2026-01-15"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["code"] == 400
    assert "批次名称不能为空" in body["message"]
    assert calls["query"] == []
    assert calls["write"] == []


def test_create_raw_import_batch_requires_edu_database(monkeypatch):
    calls = _patch_write_session(monkeypatch)
    client = _raw_auth_client(monkeypatch, edu_ds_id=None)
    r = client.post(
        "/api/v1/education/raw-score-import/batches",
        data={"batch_name": "2026届高三1月期末", "exam_time": "2026-01-15"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["code"] == 400
    assert "未登记 edu 业务库" in body["message"]
    assert calls["query"] == []
    assert calls["write"] == []


def test_list_raw_import_batches(monkeypatch):
    monkeypatch.setattr(
        "datasource.db.db.execute_sql",
        lambda *_a, **_k: (
            True,
            "ok",
            {
                "columns": ["id", "batch_name", "exam_time"],
                "rows": [(1, "2026届高三1月期末", "2026-01-15")],
            },
        ),
    )
    client = _raw_auth_client(monkeypatch)
    r = client.get("/api/v1/education/raw-score-import/batches")
    assert r.status_code == 200
    body = r.json()
    assert body["code"] == 200
    assert body["data"]["batches"] == [
        {"id": 1, "batch_name": "2026届高三1月期末", "exam_time": "2026-01-15"}
    ]


def test_list_raw_import_papers_missing_and_duplicate(monkeypatch):
    captured_sql: list[str] = []

    def _fake_execute(*args, **_k):
        sql = ""
        for arg in args:
            if isinstance(arg, str) and "SELECT" in arg.upper():
                sql = arg
                captured_sql.append(arg)
        upper = sql.upper()
        if "TB_EXAM_BATCH" in upper:
            return True, "ok", {"columns": ["id", "batch_name"], "rows": [(1, "2026届高三1月期末")]}
        if "TB_SCORE_OVERVIEW" in upper:
            return True, "ok", {
                "columns": ["row_count", "school_count", "last_write_time"],
                "rows": [(0, 0, None)],
            }
        if "TB_SCORE_DETAIL" in upper:
            return True, "ok", {
                "columns": ["exam_id", "row_count", "student_count"],
                "rows": [],
            }
        return (
            True,
            "ok",
            {
                "columns": ["id", "subject", "exam_score"],
                "rows": [
                    (11, "语文", 150),
                    (12, "数学", 150),
                    (13, "数学", 150),
                ],
            },
        )

    monkeypatch.setattr("datasource.db.db.execute_sql", _fake_execute)
    monkeypatch.setattr("src.agent.education.raw_import.execute_sql", _fake_execute)
    client = _raw_auth_client(monkeypatch)
    r = client.get("/api/v1/education/raw-score-import/papers", params={"exam_batch_id": 1})
    assert r.status_code == 200
    exam_sql = next((s for s in captured_sql if isinstance(s, str) and "FROM tb_exam" in s), "")
    assert exam_sql
    assert "subject_name" not in exam_sql
    data = r.json()["data"]
    assert len(data["papers"]) == 3
    assert "数学" in data["duplicate_subjects"]
    assert "物理" in data["missing_subjects"]
    assert "语文" not in data["missing_subjects"]
    assert "数学" not in data["missing_subjects"]
    assert data["overview"]["imported"] is False
    assert data["overview"]["row_count"] == 0


def test_list_raw_import_papers_overview_already_imported(monkeypatch):
    def _fake_execute(*args, **_k):
        sql = next((a for a in args if isinstance(a, str) and "SELECT" in a.upper()), "")
        upper = sql.upper()
        if "TB_EXAM_BATCH" in upper:
            return True, "ok", {"columns": ["id", "batch_name"], "rows": [(2, "2026届高三1月期末")]}
        if "TB_SCORE_OVERVIEW" in upper:
            return True, "ok", {
                "columns": ["row_count", "school_count", "last_write_time"],
                "rows": [(17700, 12, "2026-09-01 10:00:00")],
            }
        if "TB_SCORE_DETAIL" in upper:
            return True, "ok", {
                "columns": ["exam_id", "row_count", "student_count"],
                "rows": [],
            }
        return True, "ok", {
            "columns": ["id", "subject", "exam_score"],
            "rows": [(i, s, 150) for i, s in enumerate(_SUBJECTS, start=1)],
        }

    monkeypatch.setattr("datasource.db.db.execute_sql", _fake_execute)
    monkeypatch.setattr("src.agent.education.raw_import.execute_sql", _fake_execute)
    client = _raw_auth_client(monkeypatch)
    r = client.get("/api/v1/education/raw-score-import/papers", params={"exam_batch_id": 2})
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["overview"]["imported"] is True
    assert data["overview"]["row_count"] == 17700
    assert data["overview"]["school_count"] == 12
    assert data["missing_subjects"] == []


def test_list_raw_import_papers_school_admin_filters_overview(monkeypatch):
    captured: list[str] = []

    def _fake_execute(*args, **_k):
        sql = next((a for a in args if isinstance(a, str) and "SELECT" in a.upper()), "")
        captured.append(sql)
        upper = sql.upper()
        if "TB_EXAM_BATCH" in upper:
            return True, "ok", {"columns": ["id", "batch_name"], "rows": [(1, "2026届高三1月期末")]}
        if "TB_SCHOOL" in upper:
            return True, "ok", {"columns": ["s_name"], "rows": [("A05仪征中学",)]}
        if "TB_SCORE_OVERVIEW" in upper:
            return True, "ok", {
                "columns": ["row_count", "school_count", "last_write_time"],
                "rows": [(800, 1, None)],
            }
        if "TB_SCORE_DETAIL" in upper:
            return True, "ok", {
                "columns": ["exam_id", "row_count", "student_count"],
                "rows": [],
            }
        return True, "ok", {
            "columns": ["id", "subject", "exam_score"],
            "rows": [(i, s, 150) for i, s in enumerate(_SUBJECTS, start=1)],
        }

    monkeypatch.setattr("datasource.db.db.execute_sql", _fake_execute)
    monkeypatch.setattr("src.agent.education.raw_import.execute_sql", _fake_execute)
    client = _raw_auth_client(monkeypatch, edu_role="school_admin")
    r = client.get("/api/v1/education/raw-score-import/papers", params={"exam_batch_id": 1})
    assert r.status_code == 200
    overview_sql = next((s for s in captured if "tb_score_overview" in s.lower()), "")
    assert "A05仪征中学" in overview_sql
    assert r.json()["data"]["overview"]["row_count"] == 800


def test_list_raw_import_papers_detail_already_imported(monkeypatch):
    def _fake_execute(*args, **_k):
        sql = next((a for a in args if isinstance(a, str) and "SELECT" in a.upper()), "")
        upper = sql.upper()
        if "TB_EXAM_BATCH" in upper:
            return True, "ok", {"columns": ["id", "batch_name"], "rows": [(2, "2026届高三1月期末")]}
        if "TB_SCORE_OVERVIEW" in upper:
            return True, "ok", {
                "columns": ["row_count", "school_count", "last_write_time"],
                "rows": [(100, 1, None)],
            }
        if "TB_SCORE_DETAIL" in upper:
            return True, "ok", {
                "columns": ["exam_id", "row_count", "student_count"],
                "rows": [(9, 314432, 9248), (2, 0, 0)],
            }
        return True, "ok", {
            "columns": ["id", "subject", "exam_score"],
            "rows": [(i, s, 150) for i, s in enumerate(_SUBJECTS, start=1)],
        }

    monkeypatch.setattr("datasource.db.db.execute_sql", _fake_execute)
    monkeypatch.setattr("src.agent.education.raw_import.execute_sql", _fake_execute)
    client = _raw_auth_client(monkeypatch)
    r = client.get("/api/v1/education/raw-score-import/papers", params={"exam_batch_id": 2})
    assert r.status_code == 200
    papers = {p["subject"]: p for p in r.json()["data"]["papers"]}
    assert papers["地理"]["exam_id"] == 9
    assert papers["地理"]["detail"]["imported"] is True
    assert papers["地理"]["detail"]["row_count"] == 314432
    assert papers["地理"]["detail"]["student_count"] == 9248
    assert papers["数学"]["detail"]["imported"] is False
    assert papers["数学"]["detail"]["row_count"] == 0


def test_list_raw_import_papers_school_admin_filters_detail(monkeypatch):
    captured: list[str] = []

    def _fake_execute(*args, **_k):
        sql = next((a for a in args if isinstance(a, str) and "SELECT" in a.upper()), "")
        captured.append(sql)
        upper = sql.upper()
        if "TB_EXAM_BATCH" in upper:
            return True, "ok", {"columns": ["id", "batch_name"], "rows": [(1, "2026届高三1月期末")]}
        if "TB_SCHOOL" in upper:
            return True, "ok", {"columns": ["s_name"], "rows": [("A05仪征中学",)]}
        if "TB_SCORE_OVERVIEW" in upper:
            return True, "ok", {
                "columns": ["row_count", "school_count", "last_write_time"],
                "rows": [(800, 1, None)],
            }
        if "TB_SCORE_DETAIL" in upper:
            return True, "ok", {
                "columns": ["exam_id", "row_count", "student_count"],
                "rows": [(9, 800, 80)],
            }
        return True, "ok", {
            "columns": ["id", "subject", "exam_score"],
            "rows": [(i, s, 150) for i, s in enumerate(_SUBJECTS, start=1)],
        }

    monkeypatch.setattr("datasource.db.db.execute_sql", _fake_execute)
    monkeypatch.setattr("src.agent.education.raw_import.execute_sql", _fake_execute)
    client = _raw_auth_client(monkeypatch, edu_role="school_admin")
    r = client.get("/api/v1/education/raw-score-import/papers", params={"exam_batch_id": 1})
    assert r.status_code == 200
    detail_sql = next((s for s in captured if "tb_score_detail" in s.lower()), "")
    assert "F57E7326" in detail_sql
    assert "ESCAPE" in detail_sql.upper()
    assert "student_id" in detail_sql.lower()
    assert r.json()["data"]["papers"][8]["detail"]["imported"] is True


def test_overview_execute_scans_alerts_and_does_not_recompute(monkeypatch):
    from src.agent.education.score_import import ImportResult

    scan_kwargs: list[dict] = []
    recompute_calls: list[int] = []

    monkeypatch.setattr(
        "src.agent.education.raw_import.execute_raw_overview_import",
        lambda *_a, **_k: ImportResult(
            valid_rows=1, resolved_rows=[{"school_id": "GZ_F57E7326"}]
        ),
    )
    monkeypatch.setattr(
        "src.agent.education.alert_service.scan_alerts_after_import",
        lambda *_a, **kwargs: scan_kwargs.append(kwargs) or {"inserted": 2, "exams": 2},
    )
    monkeypatch.setattr(
        "src.agent.education.score_indicator.recompute_if_bars_exist",
        lambda *_a, **_k: recompute_calls.append(1) or {},
    )
    client = _raw_auth_client(monkeypatch)
    r = client.post(
        "/api/v1/education/raw-score-import/overview-execute",
        data={"exam_batch_id": "1"},
        files={"file": ("scores.xlsx", b"dummy-xlsx", "application/octet-stream")},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["code"] == 200
    assert body["data"]["summary"]["alert_scan"] == {"pending": True}
    assert len(scan_kwargs) == 1
    assert scan_kwargs[0]["datasource_id"] == 12
    assert scan_kwargs[0]["exam_batch_id"] == 1
    assert scan_kwargs[0]["resolved_rows"] == [{"school_id": "GZ_F57E7326"}]
    assert recompute_calls == []


def test_overview_execute_scan_failure_becomes_warning(monkeypatch):
    from src.agent.education.score_import import ImportResult

    monkeypatch.setattr(
        "src.agent.education.raw_import.execute_raw_overview_import",
        lambda *_a, **_k: ImportResult(valid_rows=1, resolved_rows=[{"school_id": "GZ_X"}]),
    )
    monkeypatch.setattr(
        "src.agent.education.alert_service.scan_alerts_after_import",
        lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("scan boom")),
    )
    client = _raw_auth_client(monkeypatch)
    r = client.post(
        "/api/v1/education/raw-score-import/overview-execute",
        data={"exam_batch_id": "1"},
        files={"file": ("scores.xlsx", b"dummy-xlsx", "application/octet-stream")},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["code"] == 200
    assert body["data"]["summary"]["alert_scan"] == {"pending": True}


def test_detail_execute_does_not_scan_alerts(monkeypatch):
    from src.agent.education.score_import import ImportResult

    scan_calls: list[int] = []
    monkeypatch.setattr(
        "src.agent.education.raw_import.execute_raw_detail_import",
        lambda *_a, **_k: ImportResult(valid_rows=1, summary={"detail_upserted": 2}),
    )
    monkeypatch.setattr(
        "src.agent.education.alert_service.scan_alerts_after_import",
        lambda *_a, **_k: scan_calls.append(1) or {},
    )
    client = _raw_auth_client(monkeypatch)
    r = client.post(
        "/api/v1/education/raw-score-import/detail-execute",
        data={"exam_batch_id": "1", "exam_id": "10"},
        files={"file": ("detail.xlsx", b"dummy-xlsx", "application/octet-stream")},
    )
    assert r.status_code == 200
    assert r.json()["code"] == 200
    assert scan_calls == []


def test_overview_execute_write_failure_is_not_validation_error(monkeypatch):
    _patch_overview_io(monkeypatch)

    def boom(*_a, **_k):
        raise RuntimeError("disk full")

    monkeypatch.setattr("src.agent.education.raw_import._upsert_dict_rows", boom)
    client = _raw_auth_client(monkeypatch)
    data = _build_overview_workbook([_two_subject_overview_row()])
    r = client.post(
        "/api/v1/education/raw-score-import/overview-execute",
        data={"exam_batch_id": "1"},
        files={"file": ("scores.xlsx", data, "application/octet-stream")},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["code"] in (400, 500)
    assert "校验未通过" not in body["message"]
    assert "写入失败" in body["message"]
    assert "整文件重导" in body["message"]


def test_detail_execute_write_failure_is_not_validation_error(monkeypatch):
    from src.agent.education.score_import import ImportErrorRow, ImportResult

    monkeypatch.setattr(
        "src.agent.education.raw_import.execute_raw_detail_import",
        lambda *_a, **_k: ImportResult(
            valid_rows=1,
            error_rows=[ImportErrorRow(row=0, field="写入", message="disk full")],
        ),
    )
    client = _raw_auth_client(monkeypatch)
    r = client.post(
        "/api/v1/education/raw-score-import/detail-execute",
        data={"exam_batch_id": "1", "exam_id": "10"},
        files={"file": ("detail.xlsx", b"dummy-xlsx", "application/octet-stream")},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["code"] in (400, 500)
    assert "校验未通过" not in body["message"]
    assert "写入失败" in body["message"]
    assert "整文件重导" in body["message"]


# 与 temp/教科院/考试/2026届-高三/2026届高三1月期末/成绩宽表.xlsx 第 1 行一致。
_REF_OVERVIEW_HEADERS = [
    "KSH",
    "XX",
    "XM",
    "ZF3M",
    "ZF4M",
    "ZF6M",
    "YW",
    "YWZW",
    "SX",
    "YY",
    "YYZW",
    "WL",
    "HX",
    "SW",
    "ZZ",
    "LS",
    "DL",
    "XH",
    "SXKG",
    "SFZH",
    "ZKCJ",
    "XKQK",
    "HXZH",
    "HXDJ",
    "SWZH",
    "SWDJ",
    "ZZZH",
    "ZZDJ",
    "DLZH",
    "DLDJ",
    "XKKM",
    "XSXZ",
    "XXLB",
    "DQ",
    "QH",
    "RY",
    "RYKG",
    "RYZW",
]


def _cell_has_border(cell) -> bool:
    b = cell.border
    return bool(b.left.style or b.right.style or b.top.style or b.bottom.style)


def test_build_raw_overview_template_parses():
    data = build_raw_overview_template()
    rows, errors, headers = _parse_overview_excel(data)
    assert not errors
    assert rows == []
    assert headers == _REF_OVERVIEW_HEADERS
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert [c.value for c in ws[1]] == _REF_OVERVIEW_HEADERS
    assert _cell_has_border(ws.cell(1, 1))
    assert _cell_has_border(ws.cell(1, len(_REF_OVERVIEW_HEADERS)))


def test_build_raw_detail_template_header_merges():
    data = build_raw_detail_template(
        subject="数学",
        questions=[
            {"question_no": "单选1", "question_score": 5.0},
            {"question_no": "12", "question_score": 5.0},
            {"question_no": "15_1", "question_score": 6.0},
        ],
    )
    rows, questions, errors, subject = _parse_detail_excel(data)
    assert not errors
    assert subject == "数学"
    assert {q["question_no"] for q in questions} >= {"单选1", "12", "15_1"}
    assert rows == []
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    row2 = [c.value for c in ws[2]]
    row3 = [c.value for c in ws[3]]
    assert str(ws.cell(1, 1).value or "").startswith("小题分(数学)")
    assert row2[:6] == ["学号", "考号", "姓名", "学校", "区域", "数学"]
    assert row3[5:8] == ["全卷", "1卷", "2卷"]
    assert "单选1（5.0分）" in row3
    assert "单选1_答案" in row3
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "A1:L1" in merged
    assert "A2:A3" in merged
    assert "B2:B3" in merged
    assert "C2:C3" in merged
    assert "D2:D3" in merged
    assert "E2:E3" in merged
    assert "F2:H2" in merged
    assert "I2:J2" in merged
    assert "K2:L2" in merged
    a3 = ws.cell(3, 1)
    assert a3.border.left.style == "thin"
    assert a3.border.right.style == "thin"
    assert a3.border.bottom.style == "thin"
    xml = zipfile.ZipFile(io.BytesIO(data)).read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert 'r="A3"' in xml
    assert 'r="B1"' in xml
    f3 = ws.cell(3, 6)
    assert f3.border.bottom.style == "thin"
    assert f3.border.left.style == "thin"
    assert _cell_has_border(ws.cell(1, 1))
    assert _cell_has_border(ws.cell(3, 9))


def test_build_raw_detail_template_uses_exam_questions():
    data = build_raw_detail_template(
        subject="地理",
        questions=[
            {"question_no": "24-2", "question_score": 4},
            {"question_no": "1", "question_score": 5.0},
        ],
    )
    rows, questions, errors, subject = _parse_detail_excel(data)
    assert not errors
    assert subject == "地理"
    assert {q["question_no"] for q in questions} == {"24-2", "1"}
    assert rows == []
    wb = load_workbook(io.BytesIO(data))
    merged = {str(r) for r in wb.active.merged_cells.ranges}
    assert "A1:J1" in merged
    assert "F2:H2" in merged
    assert "I2:J2" in merged


def test_download_raw_overview_template(monkeypatch):
    client = _raw_auth_client(monkeypatch)
    r = client.get("/api/v1/education/raw-score-import/templates/overview")
    assert r.status_code == 200
    assert "spreadsheetml" in (r.headers.get("content-type") or "")
    rows, errors, headers = _parse_overview_excel(r.content)
    assert not errors
    assert "KSH" in headers
    assert rows == []


def test_download_raw_detail_template_requires_exam(monkeypatch):
    client = _raw_auth_client(monkeypatch)
    r = client.get("/api/v1/education/raw-score-import/templates/detail")
    assert "spreadsheetml" not in (r.headers.get("content-type") or "")
    body = r.json()
    assert r.status_code in (200, 400)
    message = body.get("message") or ""
    if r.status_code == 200:
        assert body["code"] == 400
    assert "科目" in message


def test_download_raw_detail_template_with_exam(monkeypatch):
    def _fake_execute(*args, **_k):
        sql = next((a for a in args if isinstance(a, str) and "SELECT" in a.upper()), "")
        upper = sql.upper()
        if "TB_EXAM_QUESTION" in upper:
            return True, "ok", {
                "columns": ["question_no", "question_score"],
                "rows": [("24-2", 4), ("1", 5)],
            }
        if "TB_EXAM" in upper:
            return True, "ok", {
                "columns": ["id", "subject", "exam_score"],
                "rows": [(9, "地理", 100)],
            }
        return True, "ok", {"columns": [], "rows": []}

    monkeypatch.setattr("datasource.db.db.execute_sql", _fake_execute)
    monkeypatch.setattr("src.agent.education.raw_import.execute_sql", _fake_execute)
    client = _raw_auth_client(monkeypatch)
    r = client.get(
        "/api/v1/education/raw-score-import/templates/detail",
        params={"exam_id": 9},
    )
    assert r.status_code == 200
    disp = r.headers.get("content-disposition") or ""
    assert "小题分导入模板(地理).xlsx" in unquote(disp)
    _rows, questions, errors, subject = _parse_detail_excel(r.content)
    assert not errors
    assert subject == "地理"
    assert {q["question_no"] for q in questions} == {"24-2", "1"}
