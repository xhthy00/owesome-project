"""教科院原始成绩导入：匿名编码、班级解析、角色校验、宽表解析与写入。"""

from __future__ import annotations

import hashlib
import hmac
import io
import logging
import math
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from datasource.db.db import WriteDbSession, execute_sql
from datasource.service.edu_permission import EduScope
from src.agent.education.question_parser import parse_questions_from_headers
from src.agent.education.score_import import (
    ImportErrorRow,
    ImportResult,
    _in_literals,
    _normalize_exam_row,
    _pick_existing_cols,
    _rows_to_dicts,
    _schema_columns,
    _table_names,
)

_SCHOOL_CIPHER_SECRET = b"yz_edu_k1"
_SCHOOL_HEX_START = 12
_SCHOOL_HEX_LEN = 8
_REQUIRED_SUBJECTS = ("语文", "数学", "英语", "物理", "化学", "生物", "历史", "政治", "地理")
_SUBJECT_CODE_TO_NAME = {
    "YW": "语文",
    "SX": "数学",
    "YY": "英语",
    "WL": "物理",
    "HX": "化学",
    "SW": "生物",
    "LS": "历史",
    "ZZ": "政治",
    "DL": "地理",
}
_OVERVIEW_REQUIRED_COLS = {"KSH", "SFZH", "XM", "XX", "ZF6M"}
_GRADE_PREFIX_RE = re.compile(r"(高一|高二|高三|初三|初二|初一)")
_JC_YEAR_RE = re.compile(r"(\d{4})届")
_ALLOWED_RAW_IMPORT_ROLES = frozenset({"bureau_admin", "school_admin"})
# 原始成绩固定写入名为 edu 的业务库，不做成环境配置、不由前端选择。
_EDU_DB_NAME = "edu"
_EXCEL_TRAILING_DOT_ZERO = re.compile(r"^(\d+)\.0+$")
_BATCH_SIZE = 500
_PARALLEL_THRESHOLD = 800
_PARALLEL_WORKERS = 4
_SUBJECT_NAME_TO_CODE = {name: code.lower() for code, name in _SUBJECT_CODE_TO_NAME.items()}
_OVERVIEW_CORE_COLS = (
    "ksh",
    "exam_name",
    "exam_batch_id",
    "sfzh",
    "xm",
    "xx",
    "bj",
    "anon_stu_id",
)
_SCORE_WRITE_COLS = (
    "exam_id",
    "student_id",
    "school_id",
    "class",
    "score",
    "subject_name",
    "exam_score",
    "exam_time",
)
_STUDENT_WRITE_COLS = ("id", "school_id", "class", "jc")
_PREVIEW_HIDE_COLS = frozenset({"ry", "rykg", "ryzw"})
# 与 edu.tb_score_overview 列注释对齐（库中无 tb_score_view）。
_OVERVIEW_COL_COMMENTS = {
    "anon_stu_id": "学号匿名编码",
    "exam_name": "考试批次名称",
    "exam_batch_id": "考试批次ID",
    "jc": "届次",
    "ksh": "考生号",
    "sfzh": "学号",
    "xm": "姓名",
    "bj": "班级",
    "xx": "学校",
    "xh": "学校编号",
    "xsxz": "学生性质",
    "xxlb": "学校类别",
    "dq": "地区",
    "qh": "地区编号",
    "xkkm": "选考科目组合",
    "xkqk": "选科编码",
    "zf3m": "语数英总分",
    "zf4m": "语数英加首选科总分",
    "zf6m": "全科总分",
    "yw": "语文",
    "ywzw": "语文作文",
    "sx": "数学",
    "sxkg": "数学客观题",
    "yy": "英语",
    "yyzw": "英语作文",
    "ry": "日语",
    "rykg": "日语口语",
    "ryzw": "日语作文",
    "wl": "物理",
    "ls": "历史",
    "hx": "化学",
    "sw": "生物",
    "zz": "政治",
    "dl": "地理",
    "hxzh": "化学赋分",
    "hxdj": "化学等级",
    "swzh": "生物赋分",
    "swdj": "生物等级",
    "zzzh": "政治赋分",
    "zzdj": "政治等级",
    "dlzh": "地理赋分",
    "dldj": "地理等级",
    "zkcj": "中考成绩",
}
_DETAIL_WRITE_COLS = (
    "exam_id",
    "student_id",
    "question_no",
    "question_id",
    "score",
    "question_score",
    "class",
)
_DETAIL_PREVIEW_COMMENTS = {
    "exam_id": "试卷ID",
    "student_id": "学号匿名编码",
    "class": "班级",
    "ksh": "考生号",
    "sfzh": "学号",
    "xm": "姓名",
    "xx": "学校",
}
_SKIP_QUESTION_TOKENS = ("答案", "全卷", "1卷", "2卷", "合计")
_SUBJECT_TITLE_RE = re.compile(r"小题分[（(]([^)）]+)[)）]")
# 与教科院成绩宽表.xlsx 第 1 行列序一致（38 列）。
_OVERVIEW_TEMPLATE_COLS = (
    "KSH",
    "XX",
    "XM",
    "ZF3M",
    "ZF4M",
    "ZF6M",
    "YW",
    "YWZW",
    "SX",
    "YY",
    "YYZW",
    "WL",
    "HX",
    "SW",
    "ZZ",
    "LS",
    "DL",
    "XH",
    "SXKG",
    "SFZH",
    "ZKCJ",
    "XKQK",
    "HXZH",
    "HXDJ",
    "SWZH",
    "SWDJ",
    "ZZZH",
    "ZZDJ",
    "DLZH",
    "DLDJ",
    "XKKM",
    "XSXZ",
    "XXLB",
    "DQ",
    "QH",
    "RY",
    "RYKG",
    "RYZW",
)
_DETAIL_IDENTITY_HEADERS = ("学号", "考号", "姓名", "学校", "区域")
_TEMPLATE_PLACEHOLDER_ROWS = 12
_THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
_HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
_HEADER_FONT = Font(bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
logger = logging.getLogger(__name__)


@dataclass
class RawOverviewRow:
    row_num: int
    ksh: str
    sfzh: str
    xm: str
    xx: str
    scores: dict[str, float | None]
    totals: dict[str, float | None]
    others: dict[str, Any] = field(default_factory=dict)


@dataclass
class RawDetailRow:
    row_num: int
    sfzh: str
    ksh: str
    school_name: str
    scores: dict[str, float]
    xm: str = ""


def _encode_school_token(s_name: str) -> str:
    """生成学校匿名编码，与 ``scripts/import_score_overview.py`` 一致（大写 ``GZ_``）。

    不要与 ``src.agent.education.school_cipher.encode_school_name`` 混淆：
    后者前缀为小写 ``gz_``，hex 也不大写。
    """
    prefix = "GZ_"
    digest = hmac.new(_SCHOOL_CIPHER_SECRET, s_name.encode("utf-8"), hashlib.sha256).hexdigest()
    hex_part = digest[_SCHOOL_HEX_START : _SCHOOL_HEX_START + _SCHOOL_HEX_LEN].upper()
    return f"{prefix}{hex_part}"


def _generate_anon_stu_id(school_token: str, sfzh: str) -> str:
    """生成学生匿名编码：``{school_token}_{SHA256(token:sfzh)[:8].upper()}``。"""
    suffix = hashlib.sha256(f"{school_token}:{sfzh}".encode("utf-8")).hexdigest()[:8].upper()
    return f"{school_token}_{suffix}"


def _ksh_to_class_name(ksh: str, batch_name: str) -> str:
    """从考生号第 4-5 位解析班级，年级前缀取自批次名，默认高三。"""
    ksh_str = str(ksh or "").strip()
    if len(ksh_str) < 5:
        return ""
    class_no = ksh_str[3:5].lstrip("0") or "0"
    matched = _GRADE_PREFIX_RE.search(str(batch_name or ""))
    prefix = matched.group(1) if matched else "高三"
    return f"{prefix}({class_no})班"


def _normalize_id_str(value: Any) -> str:
    """把 Excel 身份证/考号单元格规范成纯字符串（去空白、去尾随 ``.0``）。"""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "nat"}:
        return ""
    matched = _EXCEL_TRAILING_DOT_ZERO.fullmatch(text)
    if matched:
        return matched.group(1)
    return text


def assert_raw_import_role_allowed(scope: EduScope) -> str | None:
    """原始成绩导入仅允许局端/校管理员；教师与学生一律拒绝。"""
    if not scope.edu_role:
        return "未配置教育角色，无法导入原始成绩"
    if scope.edu_role not in _ALLOWED_RAW_IMPORT_ROLES:
        return "仅局端管理员和校管理员可导入原始成绩"
    return None


def resolve_edu_datasource_id(session: Any, workspace_oid: int) -> int | None:
    """在当前工作空间中找到连接库名为 ``edu`` 的已登记数据源。"""
    from common.utils.aes import decrypt_conf
    from datasource.crud import crud_datasource

    rows = crud_datasource.get_datasources(session, skip=0, limit=500, oid=int(workspace_oid))
    matched: list[int] = []
    for ds in rows:
        conf = decrypt_conf(ds.configuration) if getattr(ds, "configuration", None) else {}
        if str((conf or {}).get("database") or "").strip().lower() == _EDU_DB_NAME:
            matched.append(int(ds.id))
    if not matched:
        return None
    return min(matched)


_INVALID = object()
_TOTAL_COLS = ("ZF3M", "ZF4M", "ZF6M")
_OVERVIEW_CONSUMED_COLS = (
    {"KSH", "SFZH", "XM", "XX"} | set(_SUBJECT_CODE_TO_NAME) | set(_TOTAL_COLS)
)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "nat", "<na>"}:
        return ""
    return text


def _to_float_or_invalid(value: Any) -> float | None | object:
    text = _cell_text(value)
    if not text:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return _INVALID


def _parse_overview_excel(
    file_bytes: bytes,
) -> tuple[list[RawOverviewRow], list[ImportErrorRow], list[str]]:
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, dtype=str)
    except Exception as e:  # noqa: BLE001
        return [], [ImportErrorRow(row=0, field="header", message=f"无法解析 Excel: {e}")], []
    df.columns = [str(c).strip().upper() for c in df.columns]
    headers = [str(c).strip().upper() for c in df.columns]
    missing = _OVERVIEW_REQUIRED_COLS - set(df.columns)
    if missing:
        cols = ", ".join(sorted(missing))
        return [], [ImportErrorRow(row=0, field="header", message=f"缺少必需列: {cols}")], headers

    valid: list[RawOverviewRow] = []
    errors: list[ImportErrorRow] = []
    seen_sfzh: set[str] = set()
    seen_ksh: set[str] = set()

    for offset, rec in enumerate(df.to_dict(orient="records")):
        row_num = offset + 2
        ksh = _normalize_id_str(rec.get("KSH"))
        sfzh = _normalize_id_str(rec.get("SFZH"))
        xm = _cell_text(rec.get("XM"))
        xx = _cell_text(rec.get("XX"))
        if not ksh and not sfzh and not xm and not xx:
            continue
        if not ksh or not sfzh or not xx:
            if not ksh:
                errors.append(ImportErrorRow(row=row_num, field="KSH", message="考号不能为空"))
            if not sfzh:
                errors.append(ImportErrorRow(row=row_num, field="SFZH", message="身份证号不能为空"))
            if not xx:
                errors.append(ImportErrorRow(row=row_num, field="XX", message="学校不能为空"))
            continue
        if sfzh in seen_sfzh:
            errors.append(ImportErrorRow(row=row_num, field="SFZH", message="身份证号在文件内重复"))
            continue
        if ksh in seen_ksh:
            errors.append(ImportErrorRow(row=row_num, field="KSH", message="考号在文件内重复"))
            continue
        seen_sfzh.add(sfzh)
        seen_ksh.add(ksh)

        row_bad = False
        scores: dict[str, float | None] = {}
        for code, name in _SUBJECT_CODE_TO_NAME.items():
            parsed = _to_float_or_invalid(rec.get(code))
            if parsed is _INVALID:
                errors.append(ImportErrorRow(row=row_num, field=code, message=f"{name}成绩非数值"))
                row_bad = True
            else:
                scores[name] = parsed  # type: ignore[assignment]
        totals: dict[str, float | None] = {}
        for col in _TOTAL_COLS:
            parsed = _to_float_or_invalid(rec.get(col))
            if parsed is _INVALID:
                errors.append(ImportErrorRow(row=row_num, field=col, message=f"{col}非数值"))
                row_bad = True
            else:
                totals[col] = parsed  # type: ignore[assignment]
        if row_bad:
            continue
        others: dict[str, Any] = {}
        for key, value in rec.items():
            col = str(key).strip().upper()
            if col in _OVERVIEW_CONSUMED_COLS:
                continue
            text = _cell_text(value)
            if not text:
                continue
            others[col] = text
        valid.append(
            RawOverviewRow(
                row_num=row_num,
                ksh=ksh,
                sfzh=sfzh,
                xm=xm,
                xx=xx,
                scores=scores,
                totals=totals,
                others=others,
            )
        )
    return valid, errors, headers


def _load_overview_dimensions(
    db_type: str,
    config: dict[str, Any],
    exam_batch_id: int,
    xx_values: set[str] | list[str],
) -> dict:
    tables = _table_names()
    batch_table = tables.get("exam_batch", "tb_exam_batch")
    school_table = tables.get("school", "tb_school")
    exam_table = tables.get("exam", "tb_exam")
    batch_id = int(exam_batch_id)

    ok, msg, result = execute_sql(
        db_type,
        config,
        f"SELECT id, batch_name, exam_time FROM {batch_table} WHERE id = {batch_id}",
    )
    if not ok:
        raise ValueError(msg)
    batch_rows = _rows_to_dicts(result or {})
    if not batch_rows:
        raise ValueError(f"考试批次不存在: {batch_id}")
    batch = batch_rows[0]

    schools_by_name: dict[str, Any] = {}
    names = {str(x).strip() for x in xx_values if str(x).strip()}
    if names:
        ok, msg, result = execute_sql(
            db_type,
            config,
            f"SELECT id, s_name FROM {school_table} WHERE s_name IN ({_in_literals(names)})",
        )
        if not ok:
            raise ValueError(msg)
        for row in _rows_to_dicts(result or {}):
            s_name = str(row.get("s_name") or "").strip()
            if s_name:
                schools_by_name[s_name] = str(row.get("id") or "")

    ok, msg, result = execute_sql(
        db_type,
        config,
        f"SELECT id, exam_name, subject, exam_score, exam_time "
        f"FROM {exam_table} WHERE exam_batch_id = {batch_id}",
    )
    if not ok:
        raise ValueError(msg)
    exams_by_subject: dict[str, dict] = {}
    duplicate_subjects: list[str] = []
    for row in _rows_to_dicts(result or {}):
        row = _normalize_exam_row(row)
        subject = str(row.get("subject_name") or "").strip()
        if not subject:
            continue
        if subject in exams_by_subject:
            if subject not in duplicate_subjects:
                duplicate_subjects.append(subject)
            continue
        exams_by_subject[subject] = row

    return {
        "batch": batch,
        "schools_by_name": schools_by_name,
        "exams_by_subject": exams_by_subject,
        "duplicate_subjects": duplicate_subjects,
    }


def _empty_overview_status() -> dict[str, Any]:
    return {
        "imported": False,
        "row_count": 0,
        "school_count": 0,
        "last_write_time": None,
    }


def _overview_import_status(
    db_type: str,
    config: dict[str, Any],
    exam_batch_id: int,
    scope: EduScope | None = None,
) -> dict[str, Any]:
    """查询该批次宽表是否已导入；校管理员只统计本校。"""
    tables = _table_names()
    batch_table = tables.get("exam_batch", "tb_exam_batch")
    school_table = tables.get("school", "tb_school")
    overview_table = tables.get("score_overview") or "tb_score_overview"
    batch_id = int(exam_batch_id)
    try:
        ok, _msg, result = execute_sql(
            db_type,
            config,
            f"SELECT id, batch_name FROM {batch_table} WHERE id = {batch_id}",
        )
        if not ok:
            return _empty_overview_status()
        batch_rows = _rows_to_dicts(result or {})
        if not batch_rows:
            return _empty_overview_status()
        batch_name = str(batch_rows[0].get("batch_name") or "").strip()
        if not batch_name:
            return _empty_overview_status()

        xx_filter = ""
        if scope is not None and scope.edu_role == "school_admin":
            school_token = str(scope.school_id or "").strip()
            if not school_token:
                return _empty_overview_status()
            token_sql = school_token.replace("'", "''")
            ok, _msg, school_result = execute_sql(
                db_type,
                config,
                (
                    f"SELECT s_name FROM {school_table} "
                    f"WHERE CAST(id AS TEXT) = '{token_sql}' LIMIT 1"
                ),
            )
            school_rows = _rows_to_dicts(school_result or {}) if ok else []
            s_name = str((school_rows[0].get("s_name") if school_rows else "") or "").strip()
            if not s_name:
                return _empty_overview_status()
            escaped_name = s_name.replace("'", "''")
            xx_filter = f" AND xx = '{escaped_name}'"

        name_sql = batch_name.replace("'", "''")
        ok, _msg, count_result = execute_sql(
            db_type,
            config,
            (
                "SELECT COUNT(*) AS row_count, COUNT(DISTINCT xx) AS school_count, "
                "MAX(update_time) AS last_write_time "
                f"FROM {overview_table} "
                f"WHERE (exam_batch_id = {batch_id} OR exam_name = '{name_sql}')"
                f"{xx_filter}"
            ),
        )
        if not ok:
            return _empty_overview_status()
        rows = _rows_to_dicts(count_result or {})
        rec = rows[0] if rows else {}
        row_count = int(rec.get("row_count") or 0)
        school_count = int(rec.get("school_count") or 0)
        last_write = rec.get("last_write_time")
        return {
            "imported": row_count > 0,
            "row_count": row_count,
            "school_count": school_count,
            "last_write_time": str(last_write) if last_write is not None else None,
        }
    except (TypeError, ValueError, KeyError):
        return _empty_overview_status()


def _empty_detail_status() -> dict[str, Any]:
    return {"imported": False, "row_count": 0, "student_count": 0}


def _sql_like_prefix(value: str) -> str:
    return (
        str(value)
        .replace("\\", "\\\\")
        .replace("'", "''")
        .replace("%", "\\%")
        .replace("_", "\\_")
    )


def _detail_import_status_by_exam(
    db_type: str,
    config: dict[str, Any],
    exam_ids: list[Any],
    scope: EduScope | None = None,
) -> dict[int, dict[str, Any]]:
    """按试卷统计小题分是否已导入；校管理员只统计本校学生。"""
    ids = [eid for eid in (_coerce_int_id(raw) for raw in exam_ids) if eid is not None]
    if not ids:
        return {}
    tables = _table_names()
    detail_table = tables.get("score_detail") or "tb_score_detail"
    stu_filter = ""
    if scope is not None and scope.edu_role == "school_admin":
        token = str(scope.school_id or "").strip()
        if not token:
            return {}
        stu_filter = f" AND student_id LIKE '{_sql_like_prefix(token)}\\_%' ESCAPE '\\'"
    try:
        ok, _msg, result = execute_sql(
            db_type,
            config,
            (
                "SELECT exam_id, COUNT(*) AS row_count, "
                "COUNT(DISTINCT student_id) AS student_count "
                f"FROM {detail_table} "
                f"WHERE exam_id IN ({','.join(str(i) for i in ids)})"
                f"{stu_filter} "
                "GROUP BY exam_id"
            ),
        )
        if not ok:
            return {}
        out: dict[int, dict[str, Any]] = {}
        for rec in _rows_to_dicts(result or {}):
            eid = _coerce_int_id(rec.get("exam_id"))
            if eid is None:
                continue
            row_count = int(rec.get("row_count") or 0)
            student_count = int(rec.get("student_count") or 0)
            out[eid] = {
                "imported": row_count > 0,
                "row_count": row_count,
                "student_count": student_count,
            }
        return out
    except (TypeError, ValueError, KeyError):
        return {}


def _validate_overview_rows(
    rows: list[RawOverviewRow],
    dims: dict,
    scope: EduScope | None = None,
) -> tuple[list[RawOverviewRow], list[ImportErrorRow]]:
    if scope is not None and scope.edu_role == "school_admin" and not scope.school_id:
        return [], [ImportErrorRow(row=0, field="权限", message="校管理员未配置 school_id")]

    exams_by_subject = dims.get("exams_by_subject") or {}
    duplicate_subjects = list(dims.get("duplicate_subjects") or [])
    missing = [s for s in _REQUIRED_SUBJECTS if s not in exams_by_subject]
    if missing or duplicate_subjects:
        parts: list[str] = []
        if missing:
            parts.append(f"批次缺少试卷: {', '.join(missing)}，请先补齐试卷维度")
        if duplicate_subjects:
            parts.append(f"批次存在重复试卷: {', '.join(duplicate_subjects)}")
        return [], [ImportErrorRow(row=0, field="试卷", message="；".join(parts))]

    schools_by_name = dims.get("schools_by_name") or {}
    batch = dims.get("batch") or {}
    batch_name = str(batch.get("batch_name") or "")
    valid: list[RawOverviewRow] = []
    errors: list[ImportErrorRow] = []

    for r in rows:
        row_errors: list[ImportErrorRow] = []
        if r.xx not in schools_by_name:
            row_errors.append(ImportErrorRow(row=r.row_num, field="XX", message=f"学校『{r.xx}』不存在"))
        if not _ksh_to_class_name(r.ksh, batch_name):
            row_errors.append(ImportErrorRow(row=r.row_num, field="KSH", message="考号过短"))
        for subject, score in r.scores.items():
            if score is None:
                continue
            exam = exams_by_subject.get(subject) or {}
            if score < 0:
                row_errors.append(ImportErrorRow(row=r.row_num, field=subject, message=f"{subject}负分"))
                continue
            raw_max = exam.get("exam_score")
            if raw_max is None or str(raw_max).strip() == "":
                continue
            try:
                max_score = float(raw_max)
            except (TypeError, ValueError):
                continue
            if score > max_score:
                row_errors.append(
                    ImportErrorRow(row=r.row_num, field=subject, message=f"{subject}超满分")
                )
        if row_errors:
            errors.extend(row_errors)
            continue
        valid.append(r)

    if scope is not None and scope.edu_role == "school_admin" and scope.school_id:
        foreign_names: list[str] = []
        seen: set[str] = set()
        for r in valid:
            token = schools_by_name.get(r.xx)
            if token != scope.school_id and r.xx not in seen:
                seen.add(r.xx)
                foreign_names.append(r.xx)
        if foreign_names:
            sample = foreign_names[0]
            errors.append(
                ImportErrorRow(
                    row=0,
                    field="权限",
                    message=(
                        f"文件包含非本校数据（{sample} 等 {len(foreign_names)} 所），"
                        "请先筛选为本校后再导入"
                    ),
                )
            )
            return [], errors

    return valid, errors


def _upsert_dict_rows(
    db_type: str,
    config: dict[str, Any],
    table: str,
    cols: list[str],
    conflict_cols: tuple[str, ...] | list[str],
    rows: list[dict],
) -> int:
    """将 dict 行按 cols 顺序转为元组，分块 UPSERT；≥800 行时 4 线程并行。"""
    if not rows or not cols:
        return 0
    param_rows = [tuple(r.get(c) for c in cols) for r in rows]

    def _write_chunk(chunk: list[tuple]) -> int:
        with WriteDbSession(db_type, config) as session:
            try:
                ok, msg, result = session.execute_upsert_batch(
                    table, list(cols), conflict_cols, chunk, page_size=_BATCH_SIZE
                )
                if not ok:
                    session.rollback()
                    logger.error("UPSERT %s failed: %s", table, msg)
                    raise RuntimeError(msg)
                session.commit()
                return int((result or {}).get("row_count") or len(chunk))
            except RuntimeError:
                raise
            except Exception as e:
                session.rollback()
                raise RuntimeError(str(e)) from e

    if len(param_rows) < _PARALLEL_THRESHOLD:
        return _write_chunk(param_rows)

    chunks = [param_rows[i : i + _BATCH_SIZE] for i in range(0, len(param_rows), _BATCH_SIZE)]
    workers = min(_PARALLEL_WORKERS, len(chunks))
    total = 0
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(_write_chunk, chunk) for chunk in chunks]
        for fut in as_completed(futures):
            total += fut.result()
    return total


def _filter_available(rec: dict[str, Any], available: set[str]) -> dict[str, Any]:
    if not available:
        return rec
    return {k: v for k, v in rec.items() if k in available}


def _token_and_anon(r: RawOverviewRow, dims: dict) -> tuple[str, str, str]:
    batch = dims.get("batch") or {}
    batch_name = str(batch.get("batch_name") or "")
    schools_by_name = dims.get("schools_by_name") or {}
    token = str(schools_by_name.get(r.xx) or "")
    anon = _generate_anon_stu_id(token, r.sfzh) if token else ""
    bj = _ksh_to_class_name(r.ksh, batch_name)
    return token, anon, bj


def _build_overview_write_rows(
    rows: list[RawOverviewRow],
    dims: dict,
    available_overview_cols: set[str],
) -> list[dict]:
    batch = dims.get("batch") or {}
    batch_name = str(batch.get("batch_name") or "")
    exam_batch_id = batch.get("id")
    out: list[dict] = []
    for r in rows:
        token, anon, bj = _token_and_anon(r, dims)
        rec: dict[str, Any] = {
            "ksh": r.ksh,
            "exam_name": batch_name,
            "exam_batch_id": exam_batch_id,
            "sfzh": r.sfzh,
            "xm": r.xm,
            "xx": r.xx,
            "bj": bj,
            "anon_stu_id": anon,
        }
        for subject, score in r.scores.items():
            code = _SUBJECT_NAME_TO_CODE.get(subject)
            if code:
                rec[code] = score
        rec["zf3m"] = r.totals.get("ZF3M")
        rec["zf4m"] = r.totals.get("ZF4M")
        rec["zf6m"] = r.totals.get("ZF6M")
        for key, value in r.others.items():
            rec[str(key).lower()] = value
        out.append(_filter_available(rec, available_overview_cols))
    return out


def _parse_jc_year(batch_name: str) -> str | None:
    matched = _JC_YEAR_RE.search(str(batch_name or ""))
    return matched.group(1) if matched else None


def _build_student_rows(
    rows: list[RawOverviewRow],
    dims: dict,
    available_student_cols: set[str],
) -> list[dict]:
    batch = dims.get("batch") or {}
    batch_name = str(batch.get("batch_name") or "")
    include_jc = (not available_student_cols) or ("jc" in available_student_cols)
    jc = _parse_jc_year(batch_name) if include_jc else None
    out: list[dict] = []
    seen: set[str] = set()
    for r in rows:
        token, anon, bj = _token_and_anon(r, dims)
        if not anon or anon in seen:
            continue
        seen.add(anon)
        rec: dict[str, Any] = {"id": anon, "school_id": token, "class": bj}
        if jc is not None:
            rec["jc"] = jc
        out.append(_filter_available(rec, available_student_cols))
    return out


def _build_score_rows(
    rows: list[RawOverviewRow],
    dims: dict,
    available_score_cols: set[str],
) -> list[dict]:
    exams_by_subject = dims.get("exams_by_subject") or {}
    out: list[dict] = []
    for r in rows:
        token, anon, bj = _token_and_anon(r, dims)
        for subject, score in r.scores.items():
            if score is None:
                continue
            exam = exams_by_subject.get(subject) or {}
            rec = {
                "exam_id": exam.get("id"),
                "student_id": anon,
                "school_id": token,
                "class": bj,
                "score": score,
                "subject_name": subject,
                "exam_score": exam.get("exam_score"),
                "exam_time": exam.get("exam_time"),
            }
            out.append(_filter_available(rec, available_score_cols))
    return out


def _cols_for_write(preferred: list[str] | tuple[str, ...], available: set[str]) -> list[str]:
    if available:
        return _pick_existing_cols(preferred, available)
    return list(preferred)


def _write_overview_to_db(
    db_type: str,
    config: dict[str, Any],
    rows: list[RawOverviewRow],
    dims: dict,
) -> dict[str, int]:
    tables = _table_names()
    overview_table = tables.get("score_overview") or "tb_score_overview"
    student_table = tables.get("student") or "tb_student"
    score_table = tables.get("score") or "tb_score"

    ov_available = _schema_columns(db_type, config, overview_table)
    st_available = _schema_columns(db_type, config, student_table)
    sc_available = _schema_columns(db_type, config, score_table)

    ov_rows = _build_overview_write_rows(rows, dims, ov_available)
    st_rows = _build_student_rows(rows, dims, st_available)
    sc_rows = _build_score_rows(rows, dims, sc_available)

    ov_preferred: list[str] = list(_OVERVIEW_CORE_COLS)
    ov_preferred.extend(code.lower() for code in _SUBJECT_CODE_TO_NAME)
    ov_preferred.extend(("zf3m", "zf4m", "zf6m"))
    if ov_rows:
        extras = [k for k in ov_rows[0] if k not in ov_preferred]
        ov_preferred.extend(extras)
    ov_cols = _cols_for_write(ov_preferred, ov_available)
    st_cols = _cols_for_write(_STUDENT_WRITE_COLS, st_available)
    sc_cols = _cols_for_write(_SCORE_WRITE_COLS, sc_available)

    overview_upserted = _upsert_dict_rows(
        db_type, config, overview_table, ov_cols, ("ksh", "exam_name"), ov_rows
    )
    students_upserted = _upsert_dict_rows(
        db_type, config, student_table, st_cols, ("id",), st_rows
    )
    score_upserted = _upsert_dict_rows(
        db_type, config, score_table, sc_cols, ("exam_id", "student_id"), sc_rows
    )
    return {
        "overview_upserted": int(overview_upserted),
        "students_upserted": int(students_upserted),
        "score_upserted": int(score_upserted),
    }


def _preview_column_keys(excel_headers: list[str]) -> list[str]:
    keys: list[str] = ["anon_stu_id"]
    seen = {"anon_stu_id"}
    headers = [str(h).strip().upper() for h in excel_headers]
    has_bj = "BJ" in headers

    def _add(key: str) -> None:
        if key and key not in seen:
            keys.append(key)
            seen.add(key)

    for header in headers:
        key = header.lower()
        if key in _PREVIEW_HIDE_COLS:
            continue
        _add(key)
        if key == "xx" and not has_bj:
            _add("bj")
    if "bj" not in seen:
        keys.insert(1, "bj")
    return keys


def _overview_preview_cell(r: RawOverviewRow, key: str, anon: str, bj: str) -> Any:
    if key == "anon_stu_id":
        return anon
    if key == "bj":
        return bj
    if key == "xx":
        return r.xx
    if key == "xm":
        return r.xm
    if key == "sfzh":
        return r.sfzh
    if key == "ksh":
        return r.ksh
    upper = key.upper()
    if upper in _SUBJECT_CODE_TO_NAME:
        return r.scores.get(_SUBJECT_CODE_TO_NAME[upper])
    if upper in _TOTAL_COLS:
        return r.totals.get(upper)
    return r.others.get(upper)


def _preview_columns(keys: list[str], excel_headers: list[str]) -> list[dict[str, str]]:
    original = {str(h).strip().upper().lower(): str(h).strip() for h in excel_headers}
    return [
        {"key": key, "title": _OVERVIEW_COL_COMMENTS.get(key) or original.get(key) or key}
        for key in keys
    ]


def _preview_sample_and_resolved(
    valid: list[RawOverviewRow], dims: dict, excel_headers: list[str]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, str]]]:
    keys = _preview_column_keys(excel_headers)
    columns = _preview_columns(keys, excel_headers)
    sample: list[dict[str, Any]] = []
    resolved: list[dict[str, Any]] = []
    for r in valid:
        token, anon, bj = _token_and_anon(r, dims)
        sample.append({key: _overview_preview_cell(r, key, anon, bj) for key in keys})
        resolved.append({"school_id": token})
    return sample, resolved, columns


def preview_raw_overview_import(
    file_bytes: bytes,
    exam_batch_id: int,
    scope: EduScope,
    db_type: str,
    config: dict[str, Any],
) -> ImportResult:
    rows, parse_errors, excel_headers = _parse_overview_excel(file_bytes)
    per_row_parse = sum(1 for e in parse_errors if e.row > 0)
    result = ImportResult(
        total_rows=len(rows) + per_row_parse,
        error_rows=list(parse_errors),
    )
    if any(e.row == 0 for e in parse_errors):
        return result
    try:
        dims = _load_overview_dimensions(db_type, config, exam_batch_id, {r.xx for r in rows})
    except ValueError as e:
        result.error_rows.append(ImportErrorRow(row=0, field="维度", message=str(e)))
        return result
    valid, val_errors = _validate_overview_rows(rows, dims, scope=scope)
    result.error_rows.extend(val_errors)
    result.valid_rows = len(valid)
    sample, resolved, columns = _preview_sample_and_resolved(valid, dims, excel_headers)
    result.preview_sample = sample
    result.preview_columns = columns
    result.resolved_rows = resolved  # type: ignore[assignment]
    return result


def execute_raw_overview_import(
    file_bytes: bytes,
    exam_batch_id: int,
    scope: EduScope,
    db_type: str,
    config: dict[str, Any],
) -> ImportResult:
    preview = preview_raw_overview_import(file_bytes, exam_batch_id, scope, db_type, config)
    if preview.error_rows:
        return preview
    if not preview.valid_rows:
        return preview
    rows, parse_errors, _excel_headers = _parse_overview_excel(file_bytes)
    if parse_errors:
        preview.error_rows = list(parse_errors)
        return preview
    try:
        dims = _load_overview_dimensions(db_type, config, exam_batch_id, {r.xx for r in rows})
    except ValueError as e:
        preview.error_rows.append(ImportErrorRow(row=0, field="维度", message=str(e)))
        return preview
    valid, val_errors = _validate_overview_rows(rows, dims, scope=scope)
    if val_errors:
        preview.error_rows = list(val_errors)
        preview.valid_rows = len(valid)
        return preview
    try:
        counts = _write_overview_to_db(db_type, config, valid, dims)
    except RuntimeError as e:
        logger.error("raw overview import write failed: %s", e)
        preview.error_rows.append(ImportErrorRow(row=0, field="写入", message=str(e)))
        return preview
    preview.summary = {
        "overview_upserted": int(counts.get("overview_upserted") or 0),
        "students_upserted": int(counts.get("students_upserted") or 0),
        "score_upserted": int(counts.get("score_upserted") or 0),
    }
    return preview


def _header_cell(value: Any) -> str:
    return str(value or "").strip()


def _find_header_idx(row: list[Any], candidates: list[str]) -> int | None:
    for i, cell in enumerate(row):
        text = _header_cell(cell)
        for token in candidates:
            if token in text:
                return i
    return None


def _is_skipped_score_column(question_no: str, raw_header: str) -> bool:
    blob = f"{question_no} {raw_header}"
    return any(token in blob for token in _SKIP_QUESTION_TOKENS)


def _parse_detail_excel(
    file_bytes: bytes,
) -> tuple[list[RawDetailRow], list[dict[str, Any]], list[ImportErrorRow], str]:
    """解析单个小题分文件。返回 (数据行, 保留的题目定义, 错误, 识别科目)。"""
    try:
        df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None)
    except Exception as e:  # noqa: BLE001
        return [], [], [ImportErrorRow(row=0, field="file", message=f"无法解析 Excel: {e}")], ""
    if len(df_raw) < 3:
        return [], [], [ImportErrorRow(row=0, field="file", message="小题分文件至少需 3 行表头")], ""

    title = _header_cell(df_raw.iloc[0, 0] if df_raw.shape[1] else "")
    matched = _SUBJECT_TITLE_RE.search(title)
    detected_subject = matched.group(1).strip() if matched else ""

    row2 = df_raw.iloc[1].tolist()
    row3 = df_raw.iloc[2].tolist()
    questions = parse_questions_from_headers(row3)
    kept: list[dict[str, Any]] = []
    for q in questions:
        col = q.get("col_idx")
        raw = _header_cell(row3[col] if isinstance(col, int) and col < len(row3) else "")
        if _is_skipped_score_column(str(q.get("question_no") or ""), raw):
            continue
        q = dict(q)
        q["raw_header"] = raw
        kept.append(q)

    sfzh_idx = _find_header_idx(row3, ["学号"])
    if sfzh_idx is None:
        sfzh_idx = _find_header_idx(row2, ["学号"])
    ksh_idx = _find_header_idx(row3, ["考号"])
    if ksh_idx is None:
        ksh_idx = _find_header_idx(row2, ["考号"])
    xm_idx = _find_header_idx(row3, ["姓名"])
    if xm_idx is None:
        xm_idx = _find_header_idx(row2, ["姓名"])
    school_idx = _find_header_idx(row3, ["学校"])
    if school_idx is None:
        school_idx = _find_header_idx(row2, ["学校"])
    if sfzh_idx is None:
        return [], kept, [ImportErrorRow(row=0, field="header", message="缺少『学号』列")], detected_subject

    errors: list[ImportErrorRow] = []
    rows: list[RawDetailRow] = []
    data = df_raw.iloc[3:]
    for offset, rec in enumerate(data.itertuples(index=False, name=None)):
        row_num = offset + 4

        def _at(idx: int | None) -> Any:
            if idx is None or idx >= len(rec):
                return None
            return rec[idx]

        if not any(_cell_text(v) for v in rec):
            continue
        sfzh = _normalize_id_str(_at(sfzh_idx))
        if not sfzh:
            errors.append(ImportErrorRow(row=row_num, field="学号", message="学号为空"))
            continue
        ksh = _normalize_id_str(_at(ksh_idx)) if ksh_idx is not None else ""
        xm = _cell_text(_at(xm_idx)) if xm_idx is not None else ""
        school_name = _cell_text(_at(school_idx)) if school_idx is not None else ""
        scores: dict[str, float] = {}
        row_bad = False
        for q in kept:
            col = q.get("col_idx")
            q_no = str(q.get("question_no") or "")
            parsed = _to_float_or_invalid(_at(col if isinstance(col, int) else None))
            if parsed is _INVALID:
                errors.append(ImportErrorRow(row=row_num, field=q_no, message=f"题 {q_no} 得分非数值"))
                row_bad = True
                break
            scores[q_no] = 0.0 if parsed is None else float(parsed)
        if row_bad:
            continue
        rows.append(
            RawDetailRow(
                row_num=row_num,
                sfzh=sfzh,
                ksh=ksh,
                school_name=school_name,
                scores=scores,
                xm=xm,
            )
        )
    return rows, kept, errors, detected_subject


def _sql_ok(db_type: str, config: dict, sql: str) -> list[dict]:
    ok, msg, result = execute_sql(db_type, config, sql)
    if not ok:
        raise ValueError(msg)
    return _rows_to_dicts(result or {})


def _load_detail_dimensions(
    db_type: str,
    config: dict[str, Any],
    exam_batch_id: int,
    exam_id: int,
    sfzh_values: set[str],
) -> dict:
    tables = _table_names()
    batch_table = tables.get("exam_batch", "tb_exam_batch")
    exam_table = tables.get("exam", "tb_exam")
    overview_table = tables.get("score_overview") or "tb_score_overview"
    question_table = tables.get("exam_question", "tb_exam_question")
    batch_id = int(exam_batch_id)
    eid = int(exam_id)

    batch_rows = _sql_ok(
        db_type, config, f"SELECT id, batch_name FROM {batch_table} WHERE id = {batch_id}"
    )
    if not batch_rows:
        raise ValueError(f"考试批次不存在: {batch_id}")
    batch = batch_rows[0]

    exam_rows = _sql_ok(
        db_type,
        config,
        (
            f"SELECT id, exam_name, subject, exam_score, exam_time, exam_batch_id "
            f"FROM {exam_table} WHERE id = {eid}"
        ),
    )
    if not exam_rows:
        raise ValueError(f"试卷不存在: {eid}")
    exam = _normalize_exam_row(exam_rows[0])
    try:
        exam_batch = int(exam.get("exam_batch_id"))
    except (TypeError, ValueError):
        exam_batch = None
    if exam_batch != batch_id:
        raise ValueError(f"试卷 {eid} 不属于所选批次")

    overview_by_sfzh: dict[str, dict] = {}
    if sfzh_values:
        batch_name = str(batch.get("batch_name") or "").replace("'", "''")
        sfzh_sql = _in_literals({str(s) for s in sfzh_values})
        ov_rows = _sql_ok(
            db_type,
            config,
            (
                f"SELECT sfzh, anon_stu_id, xx, bj, xm, ksh FROM {overview_table} "
                f"WHERE exam_name = '{batch_name}' AND sfzh IN ({sfzh_sql})"
            ),
        )
        for row in ov_rows:
            key = _normalize_id_str(row.get("sfzh"))
            if key:
                overview_by_sfzh[key] = row

    q_rows = _sql_ok(
        db_type,
        config,
        f"SELECT id, question_no, question_score FROM {question_table} WHERE exam_id = {eid}",
    )
    questions_by_no = {str(r.get("question_no") or "").strip(): r for r in q_rows if r.get("question_no")}
    return {
        "batch": batch,
        "exam": exam,
        "overview_by_sfzh": overview_by_sfzh,
        "questions_by_no": questions_by_no,
    }


def _validate_detail_rows(
    rows: list[RawDetailRow],
    questions: list[dict[str, Any]],
    dims: dict,
    scope: EduScope,
) -> tuple[list[dict[str, Any]], list[ImportErrorRow]]:
    errors: list[ImportErrorRow] = []
    questions_by_no = dims.get("questions_by_no") or {}
    file_nos = {str(q.get("question_no") or "") for q in questions if q.get("question_no")}
    missing = sorted(file_nos - set(questions_by_no.keys()))
    if missing:
        listed = ", ".join(missing)
        return [], [ImportErrorRow(row=0, field="question", message=f"下列题号在试卷中不存在: {listed}")]

    if scope.edu_role == "school_admin" and not scope.school_id:
        return [], [ImportErrorRow(row=0, field="权限", message="校管理员未配置 school_id")]

    valid: list[dict[str, Any]] = []
    overview_by_sfzh = dims.get("overview_by_sfzh") or {}
    for r in rows:
        overview = overview_by_sfzh.get(r.sfzh)
        if not overview:
            errors.append(
                ImportErrorRow(row=r.row_num, field="学号", message="学号不在已导入的宽表中")
            )
            continue
        anon = str(overview.get("anon_stu_id") or "").strip()
        class_name = str(overview.get("bj") or "").strip()
        school_token = _encode_school_token(str(overview.get("xx") or ""))
        if not anon:
            errors.append(
                ImportErrorRow(row=r.row_num, field="学号", message="学号不在已导入的宽表中")
            )
            continue
        score_bad = False
        for q_no, score in r.scores.items():
            q = questions_by_no.get(q_no) or {}
            full = q.get("question_score")
            try:
                full_n = float(full) if full is not None and str(full).strip() != "" else None
            except (TypeError, ValueError):
                full_n = None
            if score < 0 or (full_n is not None and score > full_n):
                errors.append(
                    ImportErrorRow(
                        row=r.row_num,
                        field=q_no,
                        message=f"题 {q_no} 得分 {score} 超过满分 {full_n}",
                    )
                )
                score_bad = True
                break
        if score_bad:
            continue
        valid.append(
            {
                "row": r,
                "anon_stu_id": anon,
                "school_token": school_token,
                "class_name": class_name,
            }
        )

    if scope.edu_role == "school_admin" and scope.school_id:
        foreign = [
            item for item in valid if item["school_token"] != scope.school_id
        ]
        if foreign:
            names: list[str] = []
            seen: set[str] = set()
            for item in foreign:
                xx = str((item["row"].school_name or getattr(item["row"], "school_name", "")) or "")
                ov = overview_by_sfzh.get(item["row"].sfzh) or {}
                label = str(ov.get("xx") or xx or item["school_token"])
                if label not in seen:
                    seen.add(label)
                    names.append(label)
            sample = names[0] if names else "未知学校"
            errors.append(
                ImportErrorRow(
                    row=0,
                    field="权限",
                    message=(
                        f"文件包含非本校数据（{sample} 等 {len(names)} 所），"
                        "请先筛选为本校后再导入"
                    ),
                )
            )
            return [], errors
    return valid, errors


def _coerce_int_id(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except (TypeError, ValueError):
        return None


def _build_detail_write_rows(resolved: list[dict[str, Any]], dims: dict) -> list[dict]:
    exam_raw = (dims.get("exam") or {}).get("id")
    exam_id = _coerce_int_id(exam_raw)
    if exam_id is None:
        exam_id = exam_raw
    questions_by_no = dims.get("questions_by_no") or {}
    out: list[dict] = []
    for item in resolved:
        r: RawDetailRow = item["row"]
        for q_no, score in r.scores.items():
            q = questions_by_no.get(q_no) or {}
            out.append(
                {
                    "exam_id": exam_id,
                    "student_id": item["anon_stu_id"],
                    "question_no": q_no,
                    "question_id": _coerce_int_id(q.get("id") or q.get("question_id")),
                    "score": score,
                    "question_score": q.get("question_score"),
                    "class": item["class_name"],
                }
            )
    return out


def _detail_preview_sample_and_columns(
    valid: list[dict[str, Any]],
    questions: list[dict[str, Any]],
    dims: dict,
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    exam_id = str((dims.get("exam") or {}).get("id") or "")
    q_nos = [str(q.get("question_no") or "") for q in questions if q.get("question_no")]
    q_titles = {
        str(q.get("question_no") or ""): str(q.get("raw_header") or q.get("question_no") or "")
        for q in questions
        if q.get("question_no")
    }
    keys = ["exam_id", "student_id", "class", "ksh", "sfzh", "xm", "xx", *q_nos]
    columns = [
        {
            "key": key,
            "title": _DETAIL_PREVIEW_COMMENTS.get(key)
            or _OVERVIEW_COL_COMMENTS.get(key)
            or q_titles.get(key)
            or key,
        }
        for key in keys
    ]
    overview_by_sfzh = dims.get("overview_by_sfzh") or {}
    sample: list[dict[str, Any]] = []
    for item in valid:
        r: RawDetailRow = item["row"]
        ov = overview_by_sfzh.get(r.sfzh) or {}
        rec: dict[str, Any] = {
            "anon_stu_id": item["anon_stu_id"],
            "exam_id": exam_id,
            "student_id": item["anon_stu_id"],
            "class": item["class_name"],
            "ksh": r.ksh or _normalize_id_str(ov.get("ksh")),
            "sfzh": r.sfzh,
            "xm": r.xm or str(ov.get("xm") or "").strip(),
            "xx": r.school_name or str(ov.get("xx") or "").strip(),
        }
        for q_no in q_nos:
            rec[q_no] = r.scores.get(q_no)
        sample.append(rec)
    return sample, columns


def preview_raw_detail_import(
    file_bytes: bytes,
    exam_batch_id: int,
    exam_id: int,
    scope: EduScope,
    db_type: str,
    config: dict[str, Any],
) -> ImportResult:
    rows, questions, parse_errors, detected_subject = _parse_detail_excel(file_bytes)
    per_row = sum(1 for e in parse_errors if e.row > 0)
    result = ImportResult(
        total_rows=len(rows) + per_row,
        error_rows=list(parse_errors),
        summary={"detected_subject": detected_subject} if detected_subject else {},
    )
    if any(e.row == 0 for e in parse_errors):
        return result
    try:
        dims = _load_detail_dimensions(
            db_type, config, exam_batch_id, exam_id, {r.sfzh for r in rows}
        )
    except ValueError as e:
        result.error_rows.append(ImportErrorRow(row=0, field="维度", message=str(e)))
        return result
    exam_subject = str((dims.get("exam") or {}).get("subject_name") or "").strip()
    if detected_subject and exam_subject and detected_subject != exam_subject:
        # ImportResult has no warnings; stash in summary for API to lift
        result.summary["subject_mismatch"] = f"文件科目「{detected_subject}」与试卷科目「{exam_subject}」不一致"
    valid, val_errors = _validate_detail_rows(rows, questions, dims, scope)
    result.error_rows.extend(val_errors)
    result.valid_rows = len(valid)
    sample, columns = _detail_preview_sample_and_columns(valid, questions, dims)
    result.preview_sample = sample
    result.preview_columns = columns
    result.resolved_rows = valid  # type: ignore[assignment]
    result.summary["students_matched"] = len(valid)
    result.summary["detail_rows"] = sum(len(item["row"].scores) for item in valid)
    return result


def execute_raw_detail_import(
    file_bytes: bytes,
    exam_batch_id: int,
    exam_id: int,
    scope: EduScope,
    db_type: str,
    config: dict[str, Any],
) -> ImportResult:
    preview = preview_raw_detail_import(
        file_bytes, exam_batch_id, exam_id, scope, db_type, config
    )
    if preview.error_rows:
        return preview
    if not preview.valid_rows:
        return preview
    rows, questions, parse_errors, _subject = _parse_detail_excel(file_bytes)
    if parse_errors:
        preview.error_rows = list(parse_errors)
        return preview
    try:
        dims = _load_detail_dimensions(
            db_type, config, exam_batch_id, exam_id, {r.sfzh for r in rows}
        )
    except ValueError as e:
        preview.error_rows.append(ImportErrorRow(row=0, field="维度", message=str(e)))
        return preview
    valid, val_errors = _validate_detail_rows(rows, questions, dims, scope)
    if val_errors:
        preview.error_rows = list(val_errors)
        preview.valid_rows = len(valid)
        return preview
    write_rows = _build_detail_write_rows(valid, dims)
    tables = _table_names()
    detail_table = tables.get("score_detail") or "tb_score_detail"
    available = _schema_columns(db_type, config, detail_table)
    cols = _cols_for_write(_DETAIL_WRITE_COLS, available)
    filtered = [_filter_available(r, set(cols) if available else set(r)) for r in write_rows]
    try:
        count = _upsert_dict_rows(
            db_type, config, detail_table, cols, ("exam_id", "student_id", "question_no"), filtered
        )
    except RuntimeError as e:
        logger.error("raw detail import write failed: %s", e)
        preview.error_rows.append(ImportErrorRow(row=0, field="写入", message=str(e)))
        return preview
    preview.summary["detail_upserted"] = int(count)
    preview.summary["students_matched"] = len(valid)
    return preview


def _workbook_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _question_header_label(question_no: str, question_score: Any) -> str:
    try:
        n = float(question_score)
    except (TypeError, ValueError):
        n = 0.0
    return f"{question_no}（{n:.1f}分）"


def _is_choice_question_no(question_no: str) -> bool:
    return question_no.startswith("单选") or question_no.startswith("多选")


def _style_template_grid(ws, *, max_row: int, max_col: int, header_rows: int) -> None:
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = _THIN_BORDER
            if row_idx <= header_rows:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = _HEADER_ALIGN


def _set_template_col_widths(ws, max_col: int, *, default: float, widths: dict[int, float]) -> None:
    for col_idx in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, default)


def _detail_score_headers(questions: list[dict[str, Any]]) -> tuple[list[str], int | None]:
    headers: list[str] = []
    paper2_offset: int | None = None
    for q in questions:
        qno = str(q.get("question_no") or "").strip()
        if not qno:
            continue
        if paper2_offset is None and not _is_choice_question_no(qno):
            paper2_offset = len(headers)
        headers.append(_question_header_label(qno, q.get("question_score")))
        if _is_choice_question_no(qno):
            headers.append(f"{qno}_答案")
    return headers, paper2_offset


def _merge_if_span(ws, *, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    if end_row < start_row or end_col < start_col:
        return
    if end_row == start_row and end_col == start_col:
        return
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=end_row,
        end_column=end_col,
    )


def _apply_detail_header_merges(ws, total_cols: int, paper2_offset: int | None) -> None:
    """表头合并对齐教科院小题分：标题整行、身份列竖向、科目盖住合计列、1卷/2卷盖住题列。"""
    _merge_if_span(ws, start_row=1, start_col=1, end_row=1, end_col=total_cols)
    for col in range(1, 6):
        _merge_if_span(ws, start_row=2, start_col=col, end_row=3, end_col=col)
    _merge_if_span(ws, start_row=2, start_col=6, end_row=2, end_col=8)
    q_start = 9
    if q_start > total_cols:
        return
    if paper2_offset is None:
        _merge_if_span(ws, start_row=2, start_col=q_start, end_row=2, end_col=total_cols)
        return
    paper2_start = q_start + paper2_offset
    _merge_if_span(ws, start_row=2, start_col=q_start, end_row=2, end_col=paper2_start - 1)
    _merge_if_span(ws, start_row=2, start_col=paper2_start, end_row=2, end_col=total_cols)


def build_raw_overview_template() -> bytes:
    """生成成绩宽表导入模板（第一张表为数据表，pandas 按 sheet 0 读取）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩宽表"
    headers = list(_OVERVIEW_TEMPLATE_COLS)
    ws.append(headers)
    max_col = len(headers)
    max_row = 1 + _TEMPLATE_PLACEHOLDER_ROWS
    _style_template_grid(ws, max_row=max_row, max_col=max_col, header_rows=1)
    _set_template_col_widths(
        ws,
        max_col,
        default=10,
        widths={1: 16, 2: 18, 3: 12, 20: 16, 31: 14},
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"
    note = wb.create_sheet("填写说明")
    note.append(["成绩宽表导入说明"])
    note.append(["1. 请使用本工作簿第一张表「成绩宽表」填写，表头 38 列及顺序须与教科院成绩宽表一致，不要改英文列名。"])
    note.append(["2. 必填：KSH（考生号）、SFZH（身份证号/学号）、XM（姓名）、XX（学校，须与系统校名完全一致）、ZF6M（全科总分）。"])
    note.append(["3. 科目列为空表示未选考，不要填非数字。转换分/等级/选考组合等列按教科院原表填写。"])
    note.append(["4. 班级由考生号第 4-5 位自动解析，无需 BJ 列。"])
    note.append(["5. 从第 2 行起填写学生数据。"])
    note.append([])
    note.append(["列名", "含义"])
    for col in _OVERVIEW_TEMPLATE_COLS:
        note.append([col, _OVERVIEW_COL_COMMENTS.get(col.lower(), "")])
    return _workbook_bytes(wb)


def build_raw_detail_template(subject: str, questions: list[dict[str, Any]]) -> bytes:
    """生成小题分导入模板。表头合并对齐教科院各科小题分。"""
    subj = str(subject or "").strip() or "未命名"
    qs = [q for q in questions if str(q.get("question_no") or "").strip()]
    if not qs:
        raise ValueError("该试卷没有题目，无法生成小题分模板")
    q_headers, paper2_offset = _detail_score_headers(qs)
    prefix = [*_DETAIL_IDENTITY_HEADERS, subj, "", ""]
    row2 = [*prefix, *([""] * len(q_headers))]
    if q_headers:
        if paper2_offset != 0:
            row2[len(_DETAIL_IDENTITY_HEADERS) + 3] = "1卷"
        if paper2_offset is not None:
            row2[len(prefix) + paper2_offset] = "2卷"
    row3 = ["", "", "", "", "", "全卷", "1卷", "2卷", *q_headers]
    total_cols = max(len(row2), len(row3), 8)
    while len(row2) < total_cols:
        row2.append("")
    while len(row3) < total_cols:
        row3.append("")
    wb = Workbook()
    ws = wb.active
    ws.title = "小题分"
    ws.append([f"小题分({subj})", *([""] * (total_cols - 1))])
    ws.append(row2)
    ws.append(row3)
    max_row = 3 + _TEMPLATE_PLACEHOLDER_ROWS
    _style_template_grid(ws, max_row=max_row, max_col=total_cols, header_rows=3)
    _apply_detail_header_merges(ws, total_cols, paper2_offset)
    _style_template_grid(ws, max_row=max_row, max_col=total_cols, header_rows=3)
    _set_template_col_widths(
        ws,
        total_cols,
        default=14,
        widths={1: 16, 2: 16, 3: 12, 4: 18, 5: 10},
    )
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A4"
    note = wb.create_sheet("填写说明")
    note.append(["小题分导入说明"])
    note.append(["1. 请先在导入页选择科目再下载，题号列按该卷真实题目生成。"])
    note.append(["2. 第一行必须是「小题分(科目)」，科目须与所选试卷一致，例如 小题分(地理)。"])
    note.append(["3. 第二行：学号、考号、姓名、学校、区域竖向合并；科目名盖住全卷/1卷/2卷合计列；1卷/2卷再盖住对应题列。"])
    note.append(["4. 第三行：全卷、1卷、2卷合计列之后是题号，格式如 单选1（5.0分）。选择题后可留「单选N_答案」列。"])
    note.append(["5. 学号填身份证号，须已出现在本批次成绩宽表中。从第 4 行起填写学生得分。"])
    note.append(["6. 导入时会跳过答案列与全卷/1卷/2卷合计列。"])
    return _workbook_bytes(wb)


def load_raw_detail_template_meta(
    db_type: str,
    config: dict[str, Any],
    exam_id: int,
) -> tuple[str, list[dict[str, Any]]]:
    """读取试卷科目与题目，供生成带真实题号的小题分模板。"""
    tables = _table_names()
    exam_table = tables.get("exam", "tb_exam")
    question_table = tables.get("exam_question", "tb_exam_question")
    eid = int(exam_id)
    exam_rows = _sql_ok(
        db_type,
        config,
        f"SELECT id, subject FROM {exam_table} WHERE id = {eid}",
    )
    if not exam_rows:
        raise ValueError(f"试卷不存在: {eid}")
    exam = _normalize_exam_row(exam_rows[0])
    subject = str(exam.get("subject_name") or exam.get("subject") or "").strip() or "未命名"
    q_rows = _sql_ok(
        db_type,
        config,
        (
            f"SELECT question_no, question_score FROM {question_table} "
            f"WHERE exam_id = {eid} ORDER BY question_no"
        ),
    )
    questions = [
        {
            "question_no": str(r.get("question_no") or "").strip(),
            "question_score": r.get("question_score"),
        }
        for r in q_rows
        if str(r.get("question_no") or "").strip()
    ]
    return subject, questions


__all__ = [
    "RawDetailRow",
    "RawOverviewRow",
    "_REQUIRED_SUBJECTS",
    "_encode_school_token",
    "_generate_anon_stu_id",
    "_ksh_to_class_name",
    "_load_overview_dimensions",
    "_overview_import_status",
    "_detail_import_status_by_exam",
    "_empty_detail_status",
    "_coerce_int_id",
    "_normalize_id_str",
    "_parse_detail_excel",
    "_parse_overview_excel",
    "_build_detail_write_rows",
    "_validate_detail_rows",
    "_validate_overview_rows",
    "assert_raw_import_role_allowed",
    "resolve_edu_datasource_id",
    "execute_raw_detail_import",
    "execute_raw_overview_import",
    "preview_raw_detail_import",
    "preview_raw_overview_import",
    "build_raw_overview_template",
    "build_raw_detail_template",
    "load_raw_detail_template_meta",
    "_OVERVIEW_TEMPLATE_COLS",
]
