"""成绩 Excel 导入：解析模板、校验维度、UPSERT 写入外部数据源。"""

from __future__ import annotations

import io
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

from datasource.db.db import WriteDbSession, execute_sql
from datasource.service.edu_permission import EduScope
from src.agent.education.schema_mapping import load_schema_from_config

ImportType = Literal["total", "detail"]

_TEMPLATE_FILES: dict[ImportType, str] = {
    "total": "脱敏成绩_仅总分.xlsx",
    "detail": "脱敏成绩_小题分明细.xlsx",
}
_SHEET_NAMES: dict[ImportType, str] = {
    "total": "成绩录入",
    "detail": "小题分明细",
}
_HEADER_ALIASES: dict[str, str] = {
    "学校编号": "school_id",
    "学校编号*": "school_id",
    "试卷编号": "exam_id",
    "试卷编号*": "exam_id",
    "试卷名称": "exam_name",
    "试卷名称*": "exam_name",
    "学号": "student_id",
    "学号*": "student_id",
    "班级": "class_name",
    "班级*": "class_name",
    "总分": "score",
    "总分*": "score",
    "题目编号": "question_id",
    "题目编号*": "question_id",
    "题号": "question_no",
    "题号*": "question_no",
    "题目满分": "question_score",
    "题目满分*": "question_score",
    "得分": "score",
    "得分*": "score",
}
_REQUIRED_TOTAL = {"school_id", "exam_id", "exam_name", "student_id", "class_name", "score"}
_REQUIRED_DETAIL = _REQUIRED_TOTAL | {"question_id", "question_no", "question_score"}
_LEGACY_REQUIRED_TOTAL = {"exam_name", "student_id", "class_name", "score"}
_LEGACY_REQUIRED_DETAIL = _LEGACY_REQUIRED_TOTAL | {"question_no"}
_BATCH_SIZE = 500
_PARALLEL_THRESHOLD = 800
_PARALLEL_WORKERS = 4
_PROJECT_ROOT = Path(__file__).resolve().parents[3]


@dataclass
class ParsedRow:
    row_num: int
    exam_name: str
    student_id: str
    class_name: str
    score: float | None = None
    school_id: str = ""
    exam_id: str = ""
    question_no: int | None = None
    question_id: str = ""
    question_score: float | None = None


@dataclass
class ResolvedRow:
    row_num: int
    exam_name: str
    student_id: str
    class_name: str
    score: float
    exam_id: str
    school_id: str
    subject_name: str
    exam_score: float | None
    exam_time: Any
    question_no: int | None = None
    question_id: str | None = None
    question_score: float | None = None
    create_student: bool = False


@dataclass
class ImportErrorRow:
    row: int
    field: str
    message: str


@dataclass
class ImportResult:
    total_rows: int = 0
    valid_rows: int = 0
    error_rows: list[ImportErrorRow] = field(default_factory=list)
    summary: dict[str, int] = field(default_factory=dict)
    preview_sample: list[dict[str, Any]] = field(default_factory=list)


def template_path(import_type: ImportType) -> Path:
    return _PROJECT_ROOT / _TEMPLATE_FILES[import_type]


def _normalize_header(cell: Any) -> str:
    return str(cell or "").strip()


def _is_empty_row(values: tuple[Any, ...]) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def _parse_score(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _parse_question_no(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _parse_id(value: Any) -> str:
    if value is None or str(value).strip() == "":
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def parse_excel(file_bytes: bytes, import_type: ImportType) -> list[ParsedRow]:
    """解析 Excel 模板，返回数据行列表。"""
    sheet_name = _SHEET_NAMES[import_type]
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"未找到工作表「{sheet_name}」，请使用官方模板")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return []

    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        key = _HEADER_ALIASES.get(_normalize_header(cell))
        if key:
            col_map[idx] = key

    mapped_keys = set(col_map.values())
    if import_type == "total":
        required = _REQUIRED_TOTAL if "school_id" in mapped_keys else _LEGACY_REQUIRED_TOTAL
    else:
        required = _REQUIRED_DETAIL if "school_id" in mapped_keys else _LEGACY_REQUIRED_DETAIL
    if not required.issubset(mapped_keys):
        wb.close()
        raise ValueError(f"表头缺少必填列，需要：{', '.join(sorted(required))}")

    parsed: list[ParsedRow] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        if _is_empty_row(row):
            continue
        data: dict[str, Any] = {}
        for col_idx, key in col_map.items():
            if col_idx < len(row):
                data[key] = row[col_idx]
        exam_name = str(data.get("exam_name") or "").strip()
        student_id = str(data.get("student_id") or "").strip()
        class_name = str(data.get("class_name") or "").strip()
        school_id = _parse_id(data.get("school_id"))
        exam_id = _parse_id(data.get("exam_id"))
        score = _parse_score(data.get("score"))
        question_no = _parse_question_no(data.get("question_no")) if import_type == "detail" else None
        question_id = _parse_id(data.get("question_id")) if import_type == "detail" else ""
        question_score = _parse_score(data.get("question_score")) if import_type == "detail" else None
        parsed.append(
            ParsedRow(
                row_num=row_idx,
                exam_name=exam_name,
                student_id=student_id,
                class_name=class_name,
                score=score,
                school_id=school_id,
                exam_id=exam_id,
                question_no=question_no,
                question_id=question_id,
                question_score=question_score,
            )
        )
    wb.close()
    return parsed


def assert_import_role_allowed(scope: EduScope) -> str | None:
    if not scope.edu_role:
        return "未配置教育角色，无法导入成绩"
    if scope.edu_role == "student":
        return "学生角色不允许导入成绩"
    return None


def _check_row_scope(scope: EduScope, class_name: str, school_id: str) -> str | None:
    if scope.edu_role == "bureau_admin":
        return None
    if scope.edu_role == "teacher":
        if scope.class_names and class_name not in scope.class_names:
            return f"班级「{class_name}」不在您的权限范围"
        return None
    if scope.edu_role == "school_admin":
        if scope.school_id and school_id and school_id != scope.school_id:
            return "学校不在您的权限范围"
        return None
    return None


def _resolve_school_id(
    scope: EduScope,
    row_school_id: str,
    student_school_id: str,
    override_school_id: str,
) -> tuple[str | None, str | None]:
    """返回 (school_id, error_message)。优先使用 Excel 中的学校编号。"""
    if row_school_id:
        sid = row_school_id.strip()
        if scope.edu_role == "school_admin" and scope.school_id and sid != scope.school_id:
            return None, "学校编号与您的权限不一致"
        if scope.edu_role == "teacher" and scope.school_id and sid != scope.school_id:
            return None, "学校编号与您的权限不一致"
        return sid, None
    if scope.edu_role in ("teacher", "school_admin"):
        sid = (scope.school_id or "").strip()
        if not sid:
            return None, "当前用户未配置 school_id"
        return sid, None
    if scope.edu_role == "bureau_admin":
        if student_school_id:
            return student_school_id, None
        if override_school_id:
            return override_school_id.strip(), None
        return None, "请指定 school_id，或确保 tb_student 含 school_id 且学生已关联学校"
    return None, "无法解析 school_id"


def _resolve_exam(
    row: ParsedRow,
    exams_by_id: dict[str, dict],
    exams_by_name: dict[str, list[dict]],
) -> tuple[dict[str, Any] | None, str | None]:
    """按试卷编号或试卷名称解析考试维度。"""
    if row.exam_id:
        exam = exams_by_id.get(row.exam_id)
        if not exam:
            return None, f"试卷编号「{row.exam_id}」不存在"
        db_name = str(exam.get("exam_name") or "").strip()
        if row.exam_name and db_name and db_name != row.exam_name.strip():
            return None, "试卷编号与试卷名称不一致"
        return exam, None
    if not row.exam_name:
        return None, "试卷名称不能为空"
    matches = exams_by_name.get(row.exam_name, [])
    if not matches:
        return None, f"试卷「{row.exam_name}」不存在"
    if len(matches) > 1:
        return None, f"试卷「{row.exam_name}」存在多条记录"
    return matches[0], None


def _rows_to_dicts(result: dict[str, Any]) -> list[dict[str, Any]]:
    cols = result.get("columns") or []
    rows = result.get("rows") or []
    return [dict(zip(cols, row)) for row in rows]


def _quote_ident(name: str, db_type: str) -> str:
    quote = '"' if db_type == "pg" else "`"
    return f"{quote}{name}{quote}"


# 导入写入仅允许这两张表（维度表只读校验，禁止 INSERT/UPDATE）
_WRITE_TABLE_ROLES = ("score", "score_detail")
_SCORE_WRITE_COLS = ("exam_id", "student_id", "school_id", "class", "score", "subject_name", "exam_score", "exam_time")
_DETAIL_WRITE_COLS = ("exam_id", "student_id", "question_no", "question_id", "score", "question_score", "class")
_SCORE_CONFLICT_COLS = ("exam_id", "student_id")
_DETAIL_CONFLICT_COLS = ("exam_id", "student_id", "question_no")


def _table_names() -> dict[str, str]:
    bundle = load_schema_from_config()
    if bundle is not None:
        return dict(bundle.mapping.tables)
    return {
        "exam": "tb_exam",
        "student": "tb_student",
        "school": "tb_school",
        "score": "tb_score",
        "score_detail": "tb_score_detail",
        "exam_question": "tb_exam_question",
    }


def _schema_columns(db_type: str, config: dict[str, Any], table: str) -> set[str]:
    from datasource.db.db import get_schema_info

    for t in get_schema_info(db_type, config):
        if t.get("name") == table:
            return {str(f.get("name") or "") for f in (t.get("fields") or []) if f.get("name")}
    return set()


def _pick_existing_cols(preferred: tuple[str, ...] | list[str], available: set[str]) -> list[str]:
    return [c for c in preferred if c in available]


_EXAM_SUBJECT_CANDIDATES = ("subject_name", "subject")


def _normalize_exam_row(row: dict[str, Any]) -> dict[str, Any]:
    """统一科目字段为 subject_name（兼容 tb_exam.subject）。"""
    out = dict(row)
    if not str(out.get("subject_name") or "").strip():
        for key in _EXAM_SUBJECT_CANDIDATES:
            val = out.get(key)
            if val is not None and str(val).strip():
                out["subject_name"] = str(val).strip()
                break
    elif not isinstance(out.get("subject_name"), str):
        out["subject_name"] = str(out.get("subject_name") or "").strip()
    return out


def _exam_select_cols(exam_cols: set[str]) -> list[str]:
    """构造 tb_exam 查询列；优先带上科目字段。"""
    preferred = ("id", "exam_name", "subject_name", "subject", "exam_score", "exam_time")
    if exam_cols:
        cols = _pick_existing_cols(preferred, exam_cols)
    else:
        # schema 探测失败时仍尝试 subject_name（用户库中该列用于回填 tb_score）
        cols = ["id", "exam_name", "subject_name", "exam_score", "exam_time"]
    if "id" not in cols:
        cols = ["id", *cols]
    return cols


def _load_dimensions(
    db_type: str,
    config: dict[str, Any],
    exam_names: set[str],
    exam_ids: set[str],
    student_ids: set[str],
    school_ids: set[str],
    exam_ids_for_questions: set[str],
    question_nos: set[int],
    question_ids: set[str],
) -> tuple[
    dict[str, list[dict]],
    dict[str, dict],
    dict[str, dict],
    dict[str, dict],
    dict[tuple[str, int], dict],
    dict[str, dict],
    bool,
]:
    """批量加载 exam / student / school / question 维度（只读，不写库）。"""
    tables = _table_names()
    exam_table = tables.get("exam", "tb_exam")
    student_table = tables.get("student", "tb_student")
    school_table = tables.get("school", "tb_school")
    question_table = tables.get("exam_question", "tb_exam_question")

    schema = execute_sql(db_type, config, "SELECT 1")
    if not schema[0]:
        raise ValueError(schema[1])

    exam_cols = _schema_columns(db_type, config, exam_table)
    student_cols = _schema_columns(db_type, config, student_table)
    student_has_school_id = "school_id" in student_cols

    exams_by_name: dict[str, list[dict]] = {}
    exams_by_id: dict[str, dict] = {}
    if exam_names or exam_ids:
        select_cols = _exam_select_cols(exam_cols)
        clauses: list[str] = []
        if exam_names and ("exam_name" in select_cols or not exam_cols or "exam_name" in exam_cols):
            if "exam_name" not in select_cols:
                select_cols.append("exam_name")
            clauses.append(f"exam_name IN ({_in_literals(exam_names)})")
        if exam_ids:
            clauses.append(f"id IN ({_in_literals(exam_ids)})")
        if not clauses:
            raise ValueError(f"表 {exam_table} 无法按试卷编号/名称查询")

        sql = f"SELECT {', '.join(select_cols)} FROM {exam_table} WHERE {' OR '.join(clauses)}"
        ok, msg, result = _execute_select_in(db_type, config, sql)
        if not ok and any(c in select_cols for c in _EXAM_SUBJECT_CANDIDATES):
            # 科目列不存在时回退，避免整批失败
            select_cols = [c for c in select_cols if c not in _EXAM_SUBJECT_CANDIDATES]
            sql = f"SELECT {', '.join(select_cols)} FROM {exam_table} WHERE {' OR '.join(clauses)}"
            ok, msg, result = _execute_select_in(db_type, config, sql)
        if not ok:
            raise ValueError(msg)
        for row in _rows_to_dicts(result):
            row = _normalize_exam_row(row)
            name = str(row.get("exam_name") or "")
            exams_by_name.setdefault(name, []).append(row)
            exams_by_id[str(row.get("id") or "")] = row

        # 若仍缺科目，按 exam_id 再查一次 subject_name
        missing_subject_ids = {
            eid for eid, ex in exams_by_id.items() if eid and not str(ex.get("subject_name") or "").strip()
        }
        if missing_subject_ids:
            for subject_col in _EXAM_SUBJECT_CANDIDATES:
                ok2, _msg2, result2 = _execute_select_in(
                    db_type,
                    config,
                    f"SELECT id, {subject_col} AS subject_name FROM {exam_table} "
                    f"WHERE id IN ({_in_literals(missing_subject_ids)})",
                )
                if not ok2:
                    continue
                for row in _rows_to_dicts(result2):
                    eid = str(row.get("id") or "")
                    subj = str(row.get("subject_name") or "").strip()
                    if eid and subj and eid in exams_by_id:
                        exams_by_id[eid]["subject_name"] = subj
                        for ex in exams_by_name.get(str(exams_by_id[eid].get("exam_name") or ""), []):
                            if str(ex.get("id") or "") == eid:
                                ex["subject_name"] = subj
                missing_subject_ids = {
                    eid
                    for eid, ex in exams_by_id.items()
                    if eid and not str(ex.get("subject_name") or "").strip()
                }
                if not missing_subject_ids:
                    break

    schools_by_id: dict[str, dict] = {}
    if school_ids:
        ok, msg, result = _execute_select_in(
            db_type, config,
            f"SELECT id FROM {school_table} WHERE id IN ({_in_literals(school_ids)})",
        )
        if not ok:
            raise ValueError(msg)
        for row in _rows_to_dicts(result):
            schools_by_id[str(row.get("id") or "")] = row

    students_by_id: dict[str, dict] = {}
    if student_ids:
        ok, msg, result = _execute_select_in(
            db_type, config,
            f"SELECT id{', school_id' if student_has_school_id else ''} "
            f"FROM {student_table} WHERE id IN ({_in_literals(student_ids)})",
        )
        if not ok:
            raise ValueError(msg)
        for row in _rows_to_dicts(result):
            students_by_id[str(row.get("id") or "")] = row

    questions_by_key: dict[tuple[str, int], dict] = {}
    questions_by_id: dict[str, dict] = {}
    question_cols = _schema_columns(db_type, config, question_table)
    q_select = _pick_existing_cols(
        ("id", "exam_id", "question_no", "question_score"),
        question_cols or {"id", "exam_id", "question_no", "question_score"},
    )
    if question_ids and q_select:
        ok, msg, result = _execute_select_in(
            db_type, config,
            f"SELECT {', '.join(q_select)} FROM {question_table} "
            f"WHERE id IN ({_in_literals(question_ids)})",
        )
        if not ok:
            raise ValueError(msg)
        for row in _rows_to_dicts(result):
            qid = str(row.get("id") or "")
            questions_by_id[qid] = row
            key = (str(row.get("exam_id") or ""), int(row.get("question_no") or 0))
            questions_by_key[key] = row
    elif exam_ids_for_questions and question_nos and q_select:
        ok, msg, result = _execute_select_in(
            db_type, config,
            f"SELECT {', '.join(q_select)} FROM {question_table} "
            f"WHERE exam_id IN ({_in_literals(exam_ids_for_questions)}) "
            f"AND question_no IN ({_in_literals(question_nos)})",
        )
        if not ok:
            raise ValueError(msg)
        for row in _rows_to_dicts(result):
            qid = str(row.get("id") or "")
            questions_by_id[qid] = row
            key = (str(row.get("exam_id") or ""), int(row.get("question_no") or 0))
            questions_by_key[key] = row

    return (
        exams_by_name,
        exams_by_id,
        students_by_id,
        schools_by_id,
        questions_by_key,
        questions_by_id,
        student_has_school_id,
    )


def _in_literals(values: set[str] | set[int]) -> str:
    parts: list[str] = []
    for v in values:
        if isinstance(v, int):
            parts.append(str(v))
        else:
            escaped = str(v).replace("'", "''")
            parts.append(f"'{escaped}'")
    return ", ".join(parts) if parts else "NULL"


def _execute_select_in(db_type: str, config: dict[str, Any], sql: str) -> tuple[bool, str, dict]:
    return execute_sql(db_type, config, sql)


def validate_and_resolve(
    rows: list[ParsedRow],
    import_type: ImportType,
    scope: EduScope,
    db_type: str,
    config: dict[str, Any],
    *,
    override_school_id: str = "",
) -> tuple[list[ResolvedRow], list[ImportErrorRow]]:
    role_err = assert_import_role_allowed(scope)
    if role_err:
        return [], [ImportErrorRow(row=0, field="权限", message=role_err)]

    exam_names = {r.exam_name for r in rows if r.exam_name}
    exam_ids = {r.exam_id for r in rows if r.exam_id}
    student_ids = {r.student_id for r in rows if r.student_id}
    school_ids = {r.school_id for r in rows if r.school_id}
    question_nos: set[int] = {r.question_no for r in rows if r.question_no is not None}
    question_ids = {r.question_id for r in rows if r.question_id}
    uses_new_template = any(r.exam_id for r in rows)

    (
        exams_by_name,
        exams_by_id,
        students_by_id,
        schools_by_id,
        questions_by_key,
        questions_by_id,
        student_has_school_id,
    ) = _load_dimensions(
        db_type,
        config,
        exam_names,
        exam_ids,
        student_ids,
        school_ids,
        set(),
        question_nos,
        question_ids,
    )

    resolved: list[ResolvedRow] = []
    errors: list[ImportErrorRow] = []

    for row in rows:
        row_errors: list[ImportErrorRow] = []
        if uses_new_template and not row.school_id:
            row_errors.append(ImportErrorRow(row.row_num, "学校编号", "不能为空"))
        if uses_new_template and not row.exam_id:
            row_errors.append(ImportErrorRow(row.row_num, "试卷编号", "不能为空"))
        if not row.exam_name:
            row_errors.append(ImportErrorRow(row.row_num, "试卷名称", "不能为空"))
        if not row.student_id:
            row_errors.append(ImportErrorRow(row.row_num, "学号", "不能为空"))
        if not row.class_name:
            row_errors.append(ImportErrorRow(row.row_num, "班级", "不能为空"))
        if row.score is None:
            row_errors.append(ImportErrorRow(row.row_num, "得分", "必须为数值"))
        elif row.score < 0:
            row_errors.append(ImportErrorRow(row.row_num, "得分", "不能为负数"))
        if import_type == "detail":
            if row.question_no is None:
                row_errors.append(ImportErrorRow(row.row_num, "题号", "必须为整数"))
            if uses_new_template and not row.question_id:
                row_errors.append(ImportErrorRow(row.row_num, "题目编号", "不能为空"))
            if uses_new_template and row.question_score is None:
                row_errors.append(ImportErrorRow(row.row_num, "题目满分", "必须为数值"))

        if row_errors:
            errors.extend(row_errors)
            continue

        exam, exam_err = _resolve_exam(row, exams_by_id, exams_by_name)
        if exam_err or not exam:
            errors.append(ImportErrorRow(row.row_num, "试卷", exam_err or "无法解析试卷"))
            continue
        exam_id = str(exam.get("id") or "")

        student = students_by_id.get(row.student_id)
        create_student = student is None
        if create_student:
            student = {}

        if row.school_id and row.school_id not in schools_by_id:
            errors.append(
                ImportErrorRow(row.row_num, "学校编号", f"学校编号「{row.school_id}」不存在")
            )
            continue

        student_school = str(student.get("school_id") or "") if student_has_school_id else ""
        school_id, school_err = _resolve_school_id(
            scope, row.school_id, student_school, override_school_id
        )
        if school_err or not school_id:
            errors.append(ImportErrorRow(row.row_num, "学校", school_err or "无法确定 school_id"))
            continue

        scope_err = _check_row_scope(scope, row.class_name, school_id)
        if scope_err:
            errors.append(ImportErrorRow(row.row_num, "权限", scope_err))
            continue

        exam_score_raw = exam.get("exam_score")
        exam_score = float(exam_score_raw) if exam_score_raw is not None else None
        if import_type == "total" and exam_score is not None and row.score is not None and row.score > exam_score:
            errors.append(
                ImportErrorRow(row.row_num, "总分", f"不能超过卷面满分 {exam_score}")
            )
            continue

        question_id: str | None = None
        question_score: float | None = None
        if import_type == "detail":
            if row.question_id:
                q = questions_by_id.get(row.question_id)
                if not q:
                    errors.append(
                        ImportErrorRow(row.row_num, "题目编号", f"题目编号「{row.question_id}」不存在")
                    )
                    continue
                if str(q.get("exam_id") or "") != exam_id:
                    errors.append(ImportErrorRow(row.row_num, "题目编号", "题目不属于该试卷"))
                    continue
                if int(q.get("question_no") or 0) != int(row.question_no or 0):
                    errors.append(ImportErrorRow(row.row_num, "题号", "题号与题目编号不一致"))
                    continue
                question_id = row.question_id
                question_score = row.question_score
                if question_score is None:
                    qs_raw = q.get("question_score")
                    question_score = float(qs_raw) if qs_raw is not None else None
            else:
                qkey = (exam_id, int(row.question_no or 0))
                q = questions_by_key.get(qkey)
                if not q:
                    errors.append(
                        ImportErrorRow(row.row_num, "题号", f"题号 {row.question_no} 在试卷中不存在")
                    )
                    continue
                question_id = str(q.get("id") or "")
                qs_raw = q.get("question_score")
                question_score = float(qs_raw) if qs_raw is not None else None
            if question_score is not None and row.score is not None and row.score > question_score:
                errors.append(
                    ImportErrorRow(row.row_num, "得分", f"不能超过该题满分 {question_score}")
                )
                continue

        resolved.append(
            ResolvedRow(
                row_num=row.row_num,
                exam_name=row.exam_name,
                student_id=row.student_id,
                class_name=row.class_name,
                score=float(row.score or 0),
                exam_id=exam_id,
                school_id=school_id,
                subject_name=str(exam.get("subject_name") or "").strip(),
                exam_score=exam_score,
                exam_time=exam.get("exam_time"),
                question_no=row.question_no,
                question_id=question_id,
                question_score=question_score,
                create_student=create_student,
            )
        )

    return resolved, errors


def _check_unique_constraint(
    db_type: str,
    config: dict[str, Any],
    table: str,
    columns: list[str],
) -> tuple[bool, str]:
    """检查表是否存在覆盖指定列的唯一约束/索引。"""
    if db_type == "pg":
        col_list = ", ".join(columns)
        sql = f"""
            SELECT 1 FROM pg_indexes
            WHERE tablename = '{table.replace("'", "''")}'
              AND indexdef ILIKE '%UNIQUE%'
              AND indexdef ILIKE '%({col_list})%'
            LIMIT 1
        """
        ok, msg, result = execute_sql(db_type, config, sql)
        if ok and result.get("row_count", 0) > 0:
            return True, ""
        return False, (
            f"表 {table} 缺少 ({', '.join(columns)}) 唯一约束，无法安全 UPSERT。"
            f"请执行：CREATE UNIQUE INDEX IF NOT EXISTS idx_{table}_{'_'.join(columns)} "
            f"ON {table} ({', '.join(columns)});"
        )
    # MySQL: check via SHOW INDEX
    ok, msg, result = execute_sql(db_type, config, f"SHOW INDEX FROM `{table}`")
    if not ok:
        return False, msg
    cols_by_key: dict[str, set[str]] = {}
    col_name_idx = (result.get("columns") or []).index("Column_name") if "Column_name" in (result.get("columns") or []) else -1
    key_name_idx = (result.get("columns") or []).index("Key_name") if "Key_name" in (result.get("columns") or []) else -1
    non_unique_idx = (result.get("columns") or []).index("Non_unique") if "Non_unique" in (result.get("columns") or []) else -1
    if col_name_idx < 0 or key_name_idx < 0:
        return True, ""  # cannot verify, allow attempt
    for row in result.get("rows") or []:
        if non_unique_idx >= 0 and row[non_unique_idx] != 0:
            continue
        key = str(row[key_name_idx])
        cols_by_key.setdefault(key, set()).add(str(row[col_name_idx]))
    wanted = set(columns)
    for cols in cols_by_key.values():
        if wanted.issubset(cols):
            return True, ""
    return False, (
        f"表 {table} 缺少 ({', '.join(columns)}) 唯一约束，无法安全 UPSERT。"
    )


def _build_upsert_sql(
    db_type: str,
    table: str,
    cols: list[str],
    conflict_cols: tuple[str, ...] | list[str],
) -> str:
    q = _quote_ident
    col_list = ", ".join(q(c, db_type) for c in cols)
    placeholders = ", ".join(["%s"] * len(cols))
    conflict_set = set(conflict_cols)
    updates = ", ".join(
        f"{q(c, db_type)} = EXCLUDED.{q(c, db_type)}" if db_type == "pg"
        else f"{q(c, db_type)} = VALUES({q(c, db_type)})"
        for c in cols if c not in conflict_set
    )
    if db_type == "pg":
        conflict = ", ".join(conflict_cols)
        return (
            f"INSERT INTO {q(table, db_type)} ({col_list}) VALUES ({placeholders}) "
            f"ON CONFLICT ({conflict}) DO UPDATE SET {updates}"
        )
    return (
        f"INSERT INTO {q(table, db_type)} ({col_list}) VALUES ({placeholders}) "
        f"ON DUPLICATE KEY UPDATE {updates}"
    )


def _build_score_upsert_sql(db_type: str, table: str, cols: list[str] | None = None) -> str:
    return _build_upsert_sql(
        db_type, table, cols or list(_SCORE_WRITE_COLS), _SCORE_CONFLICT_COLS
    )


def _build_detail_upsert_sql(db_type: str, table: str, cols: list[str] | None = None) -> str:
    return _build_upsert_sql(
        db_type, table, cols or list(_DETAIL_WRITE_COLS), _DETAIL_CONFLICT_COLS
    )


def _row_params_for_cols(r: ResolvedRow, cols: list[str]) -> tuple:
    values = {
        "exam_id": r.exam_id,
        "student_id": r.student_id,
        "school_id": r.school_id,
        "class": r.class_name,
        "score": r.score,
        "subject_name": r.subject_name,
        "exam_score": r.exam_score,
        "exam_time": r.exam_time,
        "question_no": r.question_no,
        "question_id": r.question_id,
        "question_score": r.question_score,
    }
    return tuple(values[c] for c in cols)


def _score_row_params(r: ResolvedRow, cols: list[str] | None = None) -> tuple:
    return _row_params_for_cols(r, cols or list(_SCORE_WRITE_COLS))


def _detail_row_params(r: ResolvedRow, cols: list[str] | None = None) -> tuple:
    return _row_params_for_cols(r, cols or list(_DETAIL_WRITE_COLS))


def preview_import(
    file_bytes: bytes,
    import_type: ImportType,
    scope: EduScope,
    db_type: str,
    config: dict[str, Any],
    *,
    override_school_id: str = "",
) -> ImportResult:
    try:
        parsed = parse_excel(file_bytes, import_type)
    except ValueError as e:
        return ImportResult(error_rows=[ImportErrorRow(0, "文件", str(e))])

    try:
        resolved, errors = validate_and_resolve(
            parsed, import_type, scope, db_type, config, override_school_id=override_school_id
        )
    except ValueError as e:
        return ImportResult(
            total_rows=len(parsed),
            error_rows=[ImportErrorRow(0, "校验", str(e))],
        )
    sample = [
        {
            "school_id": r.school_id,
            "exam_id": r.exam_id,
            "exam_name": r.exam_name,
            "student_id": r.student_id,
            "class_name": r.class_name,
            "score": r.score,
            **({"create_student": True} if r.create_student else {}),
            **({"question_id": r.question_id, "question_no": r.question_no, "question_score": r.question_score}
               if r.question_no is not None else {}),
        }
        for r in resolved[:10]
    ]
    students_to_create = len({r.student_id for r in resolved if r.create_student})
    return ImportResult(
        total_rows=len(parsed),
        valid_rows=len(resolved),
        error_rows=errors,
        summary={
            "inserted": 0,
            "updated": 0,
            "score_upserted": 0,
            "students_to_create": students_to_create,
            "students_created": 0,
        },
        preview_sample=sample,
    )


def _resolve_write_target(
    import_type: ImportType,
) -> tuple[str, tuple[str, ...], tuple[str, ...]]:
    """返回 (table_role, preferred_cols, conflict_cols)。成绩写入仅允许 score / score_detail。"""
    if import_type == "total":
        return "score", _SCORE_WRITE_COLS, _SCORE_CONFLICT_COLS
    if import_type == "detail":
        return "score_detail", _DETAIL_WRITE_COLS, _DETAIL_CONFLICT_COLS
    raise ValueError(f"不支持的 import_type: {import_type}")


def _ensure_missing_students(
    session: WriteDbSession,
    db_type: str,
    config: dict[str, Any],
    rows: list[ResolvedRow],
) -> tuple[int, str | None]:
    """将缺失学号写入 tb_student；返回 (新增数, 错误信息)。"""
    pending = {r.student_id: r for r in rows if r.create_student}
    if not pending:
        return 0, None

    tables = _table_names()
    student_table = tables.get("student", "tb_student")
    available = _schema_columns(db_type, config, student_table)
    preferred = ("id", "class", "school_id")
    cols = _pick_existing_cols(preferred, available) if available else ["id"]
    if "id" not in cols:
        return 0, f"表 {student_table} 缺少主键列 id，无法自动新增学生"
    if not available:
        cols = ["id"]

    param_rows: list[tuple] = []
    for student_id, row in pending.items():
        values = {
            "id": student_id,
            "class": row.class_name,
            "school_id": row.school_id,
        }
        param_rows.append(tuple(values[c] for c in cols))

    # 学生表用 ON CONFLICT DO NOTHING / INSERT IGNORE，走批量
    ok, msg, _result = session.execute_upsert_batch(
        student_table,
        cols,
        ("id",),
        param_rows,
        page_size=_BATCH_SIZE,
    )
    if not ok:
        # 批量失败的 msg 已包含真实根因（NOT NULL/外键/缺列等），直接返回。
        # 不要在同一个（已 rollback 的）session 上逐条重试：PG 下逐条同样会触犯同一约束，
        # 且旧实现会在已中止事务上继续发 SQL，把真实错误淹没在 "current transaction is aborted" 级联里。
        return 0, f"新增学生失败：{msg}"
    return len(param_rows), None


def _write_rows_batch(
    db_type: str,
    config: dict[str, Any],
    table: str,
    write_cols: list[str],
    conflict_cols: tuple[str, ...] | list[str],
    rows: list[ResolvedRow],
) -> tuple[bool, str, int, int]:
    """单连接批量写入一块数据。返回 (ok, message, row_count, first_row_num)。"""
    if not rows:
        return True, "Success", 0, 0
    first_row = rows[0].row_num
    params = [_row_params_for_cols(r, write_cols) for r in rows]
    with WriteDbSession(db_type, config) as session:
        try:
            ok, msg, result = session.execute_upsert_batch(
                table, write_cols, conflict_cols, params, page_size=_BATCH_SIZE
            )
            if not ok:
                session.rollback()
                return False, msg, 0, first_row
            session.commit()
            return True, "Success", int(result.get("row_count") or len(rows)), first_row
        except Exception as e:
            session.rollback()
            return False, str(e), 0, first_row


def _write_rows_parallel(
    db_type: str,
    config: dict[str, Any],
    table: str,
    write_cols: list[str],
    conflict_cols: tuple[str, ...] | list[str],
    rows: list[ResolvedRow],
) -> tuple[bool, str, int, int]:
    """分块多线程写入；任一块失败则返回错误。"""
    if len(rows) < _PARALLEL_THRESHOLD:
        return _write_rows_batch(db_type, config, table, write_cols, conflict_cols, rows)

    chunks = [rows[i : i + _BATCH_SIZE] for i in range(0, len(rows), _BATCH_SIZE)]
    workers = min(_PARALLEL_WORKERS, len(chunks))
    total = 0
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [
            pool.submit(
                _write_rows_batch,
                db_type,
                config,
                table,
                write_cols,
                conflict_cols,
                chunk,
            )
            for chunk in chunks
        ]
        for fut in as_completed(futures):
            ok, msg, count, first_row = fut.result()
            if not ok:
                return False, msg, total, first_row
            total += count
    return True, "Success", total, 0


def import_scores(
    file_bytes: bytes,
    import_type: ImportType,
    scope: EduScope,
    db_type: str,
    config: dict[str, Any],
    *,
    override_school_id: str = "",
) -> ImportResult:
    """执行导入。成绩写入仅限 tb_score / tb_score_detail；缺失学号会自动写入 tb_student。"""
    preview = preview_import(
        file_bytes, import_type, scope, db_type, config, override_school_id=override_school_id
    )
    if preview.error_rows:
        return preview

    tables = _table_names()
    role, preferred_cols, conflict_cols = _resolve_write_target(import_type)
    if role not in _WRITE_TABLE_ROLES:
        preview.error_rows.append(ImportErrorRow(0, "写入", "禁止写入非成绩表"))
        return preview

    target_table = tables.get(role) or ("tb_score" if role == "score" else "tb_score_detail")
    allowed_physical = {
        tables.get("score", "tb_score"),
        tables.get("score_detail", "tb_score_detail"),
        "tb_score",
        "tb_score_detail",
    }
    if target_table not in allowed_physical:
        preview.error_rows.append(
            ImportErrorRow(0, "写入", f"拒绝写入非成绩表：{target_table}")
        )
        return preview

    ok, msg = _check_unique_constraint(db_type, config, target_table, list(conflict_cols))
    if not ok:
        preview.error_rows.append(ImportErrorRow(0, "数据库", msg))
        return preview

    available = _schema_columns(db_type, config, target_table)
    write_cols = _pick_existing_cols(preferred_cols, available) if available else list(preferred_cols)
    for required in conflict_cols:
        if required not in write_cols:
            preview.error_rows.append(
                ImportErrorRow(0, "数据库", f"表 {target_table} 缺少必要列 {required}")
            )
            return preview
    if "score" not in write_cols:
        preview.error_rows.append(
            ImportErrorRow(0, "数据库", f"表 {target_table} 缺少必要列 score")
        )
        return preview

    parsed = parse_excel(file_bytes, import_type)
    try:
        resolved, errors = validate_and_resolve(
            parsed, import_type, scope, db_type, config, override_school_id=override_school_id
        )
    except ValueError as e:
        preview.error_rows.append(ImportErrorRow(0, "校验", str(e)))
        return preview
    if errors:
        preview.error_rows = errors
        preview.valid_rows = len(resolved)
        return preview

    students_created = 0
    # 先提交学生新增，再并行写成绩（各块独立连接/事务）
    with WriteDbSession(db_type, config) as session:
        try:
            created, err = _ensure_missing_students(session, db_type, config, resolved)
            if err:
                session.rollback()
                preview.error_rows.append(ImportErrorRow(0, "学生", err))
                return preview
            students_created = created
            session.commit()
        except Exception as e:
            session.rollback()
            preview.error_rows.append(ImportErrorRow(0, "学生", str(e)))
            return preview

    ok, msg, upserted, first_row = _write_rows_parallel(
        db_type, config, target_table, write_cols, conflict_cols, resolved
    )
    if not ok:
        preview.error_rows.append(ImportErrorRow(first_row or 0, "写入", msg))
        return preview

    preview.summary = {
        "inserted": 0,
        "updated": upserted,
        "score_upserted": upserted if import_type == "total" else 0,
        "students_to_create": len({r.student_id for r in resolved if r.create_student}),
        "students_created": students_created,
    }
    return preview


def import_result_to_dict(result: ImportResult) -> dict[str, Any]:
    return {
        "total_rows": result.total_rows,
        "valid_rows": result.valid_rows,
        "error_rows": [
            {"row": e.row, "field": e.field, "message": e.message}
            for e in result.error_rows
        ],
        "summary": result.summary,
        "preview_sample": result.preview_sample,
    }
